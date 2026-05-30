
from __future__ import annotations

import json
import argparse
import datetime as dt
import re
import time
import os
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# ==============================
# GAME EDGE ENGINE V1 - MERGED
# Adds projected win %, volatility label, parlay grade, and ML recommendation to Game Rankings.
# ==============================

def _gee_safe_float(value, default=0.0):
    try:
        if value is None or pd.isna(value):
            return default
        return float(value)
    except Exception:
        return default

def gee_american_odds_to_implied_pct(odds):
    odds = _gee_safe_float(odds, None)
    if odds is None or odds == 0:
        return None
    if odds < 0:
        return round(abs(odds) / (abs(odds) + 100) * 100, 1)
    return round(100 / (odds + 100) * 100, 1)

def gee_edge_to_projected_win_pct(row):
    edge = _gee_safe_float(row.get("edge_vs_opponent"))
    offense_adv = _gee_safe_float(row.get("offense_advantage"))
    pitcher_adj = _gee_safe_float(row.get("pitcher_score_adj"))
    opp_pitcher_adj = _gee_safe_float(row.get("opponent_pitcher_score_adj"))
    volatility_penalty = _gee_safe_float(row.get("volatility_penalty_ml"))
    public_penalty = _gee_safe_float(row.get("public_penalty_ml"))
    pitcher_gap = pitcher_adj - opp_pitcher_adj
    win_pct = 50.0
    win_pct += max(min(edge * 0.45, 12), -12)
    win_pct += max(min(pitcher_gap * 0.07, 5), -5)
    win_pct += max(min(offense_adv * 4.0, 5), -5)
    win_pct -= volatility_penalty * 6
    win_pct -= public_penalty * 5
    short_leash = str(row.get("short_leash_flag", "")).lower()
    if "yes" in short_leash:
        win_pct -= 2.5
    opp_type = str(row.get("opponent_pitcher_pick_type", "")).lower()
    if "strong sp" in opp_type:
        win_pct -= 2.0
    win_pct = max(38.0, min(66.0, win_pct))
    return round(win_pct, 1)

def gee_get_volatility_label(row):
    volatility = _gee_safe_float(row.get("team_volatility"), 1.0)
    volatility_penalty = _gee_safe_float(row.get("volatility_penalty_ml"), 0.0)
    if volatility >= 1.18 or volatility_penalty >= 0.35:
        return "High"
    if volatility >= 1.08 or volatility_penalty >= 0.15:
        return "Medium"
    return "Low"

def gee_get_parlay_grade(projected_win_pct, volatility_label, model_edge_pct=None):
    edge = _gee_safe_float(model_edge_pct, 0)
    # If no Vegas odds are available, grade off win% + volatility only.
    if model_edge_pct is None:
        if projected_win_pct >= 65 and volatility_label == "Low":
            return "A"
        if projected_win_pct >= 62 and volatility_label in ["Low", "Medium"]:
            return "B"
        if projected_win_pct >= 58 and volatility_label != "High":
            return "C"
        return "Pass"
    if projected_win_pct >= 65 and volatility_label == "Low" and edge >= 3:
        return "A"
    if projected_win_pct >= 62 and volatility_label in ["Low", "Medium"] and edge >= 1:
        return "B"
    if projected_win_pct >= 58 and volatility_label != "High":
        return "C"
    return "Pass"

def gee_get_ml_recommendation(projected_win_pct, parlay_grade, volatility_label):
    if parlay_grade == "A":
        return "Elite ML / Parlay Anchor"
    if parlay_grade == "B":
        return "Strong ML Lean"
    if parlay_grade == "C":
        return "Small Edge / Research Only"
    if volatility_label == "High":
        return "Avoid - High Volatility"
    return "Avoid"

def add_game_edge_engine(game_rankings_df, odds_col="moneyline_odds"):
    if game_rankings_df is None or game_rankings_df.empty:
        return game_rankings_df
    df = game_rankings_df.copy()
    df["projected_win_pct"] = df.apply(gee_edge_to_projected_win_pct, axis=1)
    df["volatility_label"] = df.apply(gee_get_volatility_label, axis=1)
    if odds_col in df.columns:
        df["vegas_implied_pct"] = df[odds_col].apply(gee_american_odds_to_implied_pct)
    else:
        df["vegas_implied_pct"] = None
    def _calc_edge(row):
        vegas = row.get("vegas_implied_pct")
        if vegas is None:
            return None
        return round(_gee_safe_float(row.get("projected_win_pct")) - _gee_safe_float(vegas), 1)
    df["model_edge_pct"] = df.apply(_calc_edge, axis=1)
    df["parlay_grade"] = df.apply(lambda r: gee_get_parlay_grade(_gee_safe_float(r.get("projected_win_pct")), r.get("volatility_label"), r.get("model_edge_pct")), axis=1)
    df["ml_recommendation_v2"] = df.apply(lambda r: gee_get_ml_recommendation(_gee_safe_float(r.get("projected_win_pct")), r.get("parlay_grade"), r.get("volatility_label")), axis=1)
    return df




def _clean_value(v):
    """Convert pandas/numpy values into JSON-safe Python values."""
    try:
        import math
        import numpy as np
        import pandas as pd
    except Exception:
        np = None
        pd = None
        math = None

    if v is None:
        return None

    if pd is not None and pd.isna(v):
        return None

    if np is not None:
        if isinstance(v, (np.integer,)):
            return int(v)
        if isinstance(v, (np.floating,)):
            return None if np.isnan(v) else float(v)
        if isinstance(v, (np.bool_,)):
            return bool(v)

    if isinstance(v, float):
        if math is not None and math.isnan(v):
            return None
        return float(v)

    if isinstance(v, (int, str, bool)):
        return v

    return str(v)


def df_to_records(df):
    """Convert dataframe to JSON-safe list of dicts."""
    if df is None or len(df) == 0:
        return []

    records = df.to_dict(orient="records")
    cleaned = []
    for row in records:
        cleaned.append({k: _clean_value(v) for k, v in row.items()})
    return cleaned


def build_final_card_json(final_card_df):
    """Layer 1: final betting card."""
    return {
        "generated_section": "final_card",
        "plays": df_to_records(final_card_df)
    }


def build_game_cards_json(player_rows, game_rankings, pitcher_line_value):
    """
    Layer 2: group picks by game.
    Each game gets:
    - ml_lean
    - top 2 HR picks
    - top 2 hit picks
    - top K prop if available
    """
    games_output = []

    if game_rankings is None or len(game_rankings) == 0:
        return games_output

    gr = game_rankings.copy()
    pr = player_rows.copy() if player_rows is not None else None
    plv = pitcher_line_value.copy() if pitcher_line_value is not None else None

    unique_games = gr["game"].dropna().unique().tolist()

    for game_name in unique_games:
        game_rank = gr[gr["game"] == game_name].copy()

        ml_lean = None
        if len(game_rank) > 0:
            best_ml_row = game_rank.sort_values("edge_vs_opponent", ascending=False).iloc[0]
            ml_lean = {
                "team": _clean_value(best_ml_row.get("teamName")),
                "opponent": _clean_value(best_ml_row.get("opponentTeam")),
                "edge_vs_opponent": _clean_value(best_ml_row.get("edge_vs_opponent")),
                "recommended_play": _clean_value(best_ml_row.get("recommended_play")),
                "pitcher_pick_type": _clean_value(best_ml_row.get("pitcher_pick_type")),
                "opponent_pitcher_pick_type": _clean_value(best_ml_row.get("opponent_pitcher_pick_type")),
            }

        game_hit_picks = []
        game_hr_picks = []
        game_k_pick = None

        if pr is not None and len(pr) > 0:
            game_players = pr[pr["game"] == game_name].copy() if "game" in pr.columns else pd.DataFrame()

            if len(game_players) > 0:
                if "Hit_score" in game_players.columns:
                    top_hits = game_players.sort_values("Hit_score", ascending=False).head(2)
                    hit_cols = [c for c in [
                        "playerName", "teamName", "opponent_pitcher",
                        "opponent_pitcher_pick_type", "Hit_score",
                        "lineup_status", "batting_order_slot", "park_favorability"
                    ] if c in top_hits.columns]
                    game_hit_picks = df_to_records(top_hits[hit_cols])

                if "HR_score" in game_players.columns:
                    top_hrs = game_players.sort_values("HR_score", ascending=False).head(2)
                    hr_cols = [c for c in [
                        "playerName", "teamName", "opponent_pitcher",
                        "opponent_pitcher_pick_type", "HR_score",
                        "lineup_status", "batting_order_slot", "park_favorability"
                    ] if c in top_hrs.columns]
                    game_hr_picks = df_to_records(top_hrs[hr_cols])

        if plv is not None and len(plv) > 0 and len(game_rank) > 0:
            teams_in_game = set(game_rank["teamName"].dropna().tolist())
            game_pitchers = plv[plv["teamName"].isin(teams_in_game)].copy()

            if len(game_pitchers) > 0:
                sort_col = "projected_k_mid" if "projected_k_mid" in game_pitchers.columns else "pitcher_score_adj"
                game_pitchers = game_pitchers.sort_values(sort_col, ascending=False)
                best_k = game_pitchers.iloc[0]
                game_k_pick = {
                    "pitcherName": _clean_value(best_k.get("pitcherName")),
                    "teamName": _clean_value(best_k.get("teamName")),
                    "opponentTeam": _clean_value(best_k.get("opponentTeam")),
                    "recommended_k_action": _clean_value(best_k.get("recommended_k_action")),
                    "max_playable_k_line": _clean_value(best_k.get("max_playable_k_line")),
                    "projected_k_floor": _clean_value(best_k.get("projected_k_floor")),
                    "projected_k_mid": _clean_value(best_k.get("projected_k_mid")),
                    "projected_k_ceiling": _clean_value(best_k.get("projected_k_ceiling")),
                    "pick_type": _clean_value(best_k.get("pick_type")),
                }

        games_output.append({
            "game": game_name,
            "ml_lean": ml_lean,
            "top_hit_picks": game_hit_picks,
            "top_hr_picks": game_hr_picks,
            "top_k_pick": game_k_pick
        })

    return games_output


def build_research_json(
    game_rankings,
    pitcher_metrics,
    pitcher_line_value,
    hr_drought,
    hit_drought,
    top_picks,
    refined_picks,
    final_card_df,
    plus_money_props=None,
    hr_value_watch=None
):
    """Layer 3: all research tabs for app browsing."""
    return {
        "game_rankings": df_to_records(game_rankings),
        "pitcher_metrics": df_to_records(pitcher_metrics),
        "pitcher_line_value": df_to_records(pitcher_line_value),
        "hr_drought": df_to_records(hr_drought),
        "hit_drought": df_to_records(hit_drought),
        "top_picks": df_to_records(top_picks),
        "refined_picks": df_to_records(refined_picks),
        "plus_money_props": df_to_records(plus_money_props),
        "hr_value_watch": df_to_records(hr_value_watch),
        "final_card": df_to_records(final_card_df),
    }


def build_app_payload(
    target_date,
    final_card_df,
    player_rows,
    game_rankings,
    pitcher_metrics,
    pitcher_line_value,
    hr_drought,
    hit_drought,
    top_picks,
    refined_picks,
    plus_money_props=None,
    hr_value_watch=None
):
    """Full JSON payload for the future iPhone app."""
    return {
        "date": str(target_date),
        "final_card": build_final_card_json(final_card_df),
        "games": build_game_cards_json(player_rows, game_rankings, pitcher_line_value),
        "research": build_research_json(
            game_rankings=game_rankings,
            pitcher_metrics=pitcher_metrics,
            pitcher_line_value=pitcher_line_value,
            hr_drought=hr_drought,
            hit_drought=hit_drought,
            top_picks=top_picks,
            refined_picks=refined_picks,
            final_card_df=final_card_df,
            plus_money_props=plus_money_props,
            hr_value_watch=hr_value_watch
        )
    }


def save_app_json(payload, output_path):
    """Write JSON file to disk."""
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)


def _recent_cash_norm(value):
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def _infer_cash_bet_type(row: dict) -> str:
    """Map model rows to the graded bet_type used in results_history_latest.json."""
    bet = str(row.get("bet_type") or "").strip()
    if bet and bet.lower() not in {"", "nan", "none", "—"}:
        return bet
    typ = str(row.get("type") or row.get("category") or "").strip().upper()
    if typ == "HIT" or "HIT" in typ:
        return "1+ Hit"
    if typ == "HR" or "HOME" in typ:
        return "HR"
    return bet or typ


def _load_recent_cash_history(max_rows: int = 1000) -> list[dict]:
    """Load graded result rows so research tables can show recent cash rates.

    The results file is maintained by app_server_mobile_cloud_ready.py. This helper is read-only.
    """
    try:
        path = OUTPUT_DIR / "history" / "results_history_latest.json"
        if not path.exists():
            return []
        raw = json.loads(path.read_text(encoding="utf-8"))
        rows = raw.get("rows", []) if isinstance(raw, dict) else raw
        if not isinstance(rows, list):
            return []
        out = []
        for r in rows[:max_rows]:
            if not isinstance(r, dict):
                continue
            if r.get("result_status") not in {"Win", "Loss"}:
                continue
            pick = r.get("pick") or r.get("playerName") or r.get("pitcherName")
            bet = r.get("bet_type")
            if not pick or not bet:
                continue
            out.append(r)
        return out
    except Exception:
        return []


def add_recent_cash_rate_columns(df: pd.DataFrame, history_rows: list[dict] | None = None, default_bet_type: str | None = None, window: int = 10) -> pd.DataFrame:
    """Add recent result history to research/top-pick tables.

    Columns added:
    - recent_cash_rate: win percentage over the player's last N graded results for that bet type
    - recent_cash_record: W-L record over that same window
    - recent_cash_sample: number of graded results found in the window
    - recent_cash_last_10: compact W/L string, newest first
    """
    if df is None or df.empty:
        return df
    out = df.copy()
    rows = history_rows if history_rows is not None else _load_recent_cash_history()
    if not rows:
        out["recent_cash_rate"] = None
        out["recent_cash_record"] = "—"
        out["recent_cash_sample"] = 0
        out["recent_cash_last_10"] = "—"
        return out

    by_key = {}
    for r in rows:
        pick_key = _recent_cash_norm(r.get("pick") or r.get("playerName") or r.get("pitcherName"))
        bet_key = str(r.get("bet_type") or "").strip().lower()
        if not pick_key or not bet_key:
            continue
        by_key.setdefault((pick_key, bet_key), []).append(r)

    rates, records, samples, strings = [], [], [], []
    for _, row in out.iterrows():
        pick = row.get("playerName") or row.get("pick") or row.get("pitcherName")
        bet = default_bet_type or _infer_cash_bet_type(row.to_dict())
        key = (_recent_cash_norm(pick), str(bet or "").strip().lower())
        recent = by_key.get(key, [])[:window]
        wins = sum(1 for r in recent if r.get("result_status") == "Win")
        losses = sum(1 for r in recent if r.get("result_status") == "Loss")
        n = wins + losses
        if n:
            rates.append(round(wins / n, 3))
            records.append(f"{wins}-{losses}")
            samples.append(n)
            strings.append("".join("W" if r.get("result_status") == "Win" else "L" for r in recent))
        else:
            rates.append(None)
            records.append("—")
            samples.append(0)
            strings.append("—")
    out["recent_cash_rate"] = rates
    out["recent_cash_record"] = records
    out["recent_cash_sample"] = samples
    out["recent_cash_last_10"] = strings
    return out


def _confidence_existing_value(value):
    if value is None:
        return ""
    s = str(value).strip()
    return "" if s in {"", "—", "None", "nan", "NaN"} else s


def _rate_to_unit(value):
    try:
        if value is None or pd.isna(value):
            return None
        n = float(value)
        return n / 100.0 if n > 1.0 else n
    except Exception:
        return None


def assign_refined_pick_confidence(df: pd.DataFrame) -> pd.DataFrame:
    """Assign A+/A/B/C to Refined Picks so Results no longer groups them as Unknown.

    Uses model score first, then recent result history/current form as confirmation.
    This does not remove any picks; it only labels them for performance tracking.
    """
    if df is None or df.empty:
        return df
    out = df.copy()

    for c in ["Hit_score", "contact_quality_score", "hit_pct_last_10", "recent_cash_rate"]:
        if c not in out.columns:
            out[c] = None

    def _grade(row):
        existing = _confidence_existing_value(row.get("confidence"))
        if existing:
            return existing
        if str(row.get("category") or "").lower() == "info" or str(row.get("bet_type") or "").lower() == "no plays":
            return "Research"

        score = nz(row.get("Hit_score"), 0)
        contact = nz(row.get("contact_quality_score"), 0)
        cash = _rate_to_unit(row.get("recent_cash_rate"))
        l10 = _rate_to_unit(row.get("hit_pct_last_10"))
        sample = nz(row.get("recent_cash_sample"), 0)

        # Recent cash gets priority once there is enough sample. Otherwise use current hit form.
        support = cash if sample >= 3 and cash is not None else l10
        support = support if support is not None else 0

        slot = nz(row.get("batting_order_slot"), 99)
        confirmed = str(row.get("lineup_status") or "").eq("Confirmed Starter") if hasattr(str(row.get("lineup_status") or ""), "eq") else (str(row.get("lineup_status") or "") == "Confirmed Starter")
        opp_type = str(row.get("opponent_pitcher_pick_type") or "Neutral")

        # Stricter confidence buckets based on the actual results audit:
        # A+ must be elite, but not so tight that every strong slate produces zero official plays.
        if (
            confirmed
            and slot <= 5
            and opp_type != "Strong SP"
            and score >= 5.00
            and contact >= 3.40
            and support >= 0.80
        ):
            return "A+"
        if (
            confirmed
            and slot <= 6
            and opp_type != "Strong SP"
            and score >= 4.75
            and contact >= 2.75
            and support >= 0.70
        ):
            return "A"
        if score >= 4.20 or support >= 0.60:
            return "B"
        return "C"

    out["confidence"] = out.apply(_grade, axis=1)
    return out


DEFAULT_SEASON = 2026

def resolve_storage_dir() -> Path:
    configured = os.getenv("HR_APP_DATA_DIR")
    if configured:
        p = Path(configured)
        p.mkdir(parents=True, exist_ok=True)
        return p

    render_default = Path("/var/data/hr-picks/output")
    if render_default.parent.exists():
        render_default.mkdir(parents=True, exist_ok=True)
        return render_default

    local_default = Path("output")
    local_default.mkdir(parents=True, exist_ok=True)
    return local_default

OUTPUT_DIR = resolve_storage_dir()
SLEEP_BETWEEN_CALLS = 0.02

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
RED = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

PARK_FAVORABILITY_MAP = {
    "Athletics": "Neutral", "Cincinnati Reds": "Favorable", "New York Yankees": "Favorable",
    "Los Angeles Dodgers": "Favorable", "Los Angeles Angels": "Favorable", "Atlanta Braves": "Favorable",
    "Texas Rangers": "Favorable", "Philadelphia Phillies": "Favorable", "New York Mets": "Favorable",
    "Minnesota Twins": "Favorable", "Chicago Cubs": "Neutral", "Seattle Mariners": "Unfavorable",
    "Kansas City Royals": "Unfavorable", "Cleveland Guardians": "Unfavorable", "Detroit Tigers": "Unfavorable",
    "Tampa Bay Rays": "Unfavorable", "Oakland Athletics": "Unfavorable", "Baltimore Orioles": "Unfavorable",
    "San Francisco Giants": "Unfavorable", "Milwaukee Brewers": "Unfavorable", "Miami Marlins": "Neutral",
    "Houston Astros": "Neutral", "Toronto Blue Jays": "Neutral", "Boston Red Sox": "Neutral",
    "Washington Nationals": "Neutral", "Chicago White Sox": "Neutral", "San Diego Padres": "Neutral",
    "Pittsburgh Pirates": "Neutral", "Arizona Diamondbacks": "Neutral", "St. Louis Cardinals": "Neutral",
    "Colorado Rockies": "Favorable",
}

TEAM_VOLATILITY_MAP = {
    "Seattle Mariners": 1.25,
    "Minnesota Twins": 1.10,
    "Tampa Bay Rays": 1.10,
    "Cincinnati Reds": 1.08,
    "New York Yankees": 1.05,
}

PUBLIC_BIAS_MAP = {
    "New York Yankees": 1.15,
    "Los Angeles Dodgers": 1.12,
    "New York Mets": 1.06,
    "Atlanta Braves": 1.05,
}

MAX_REFINED_PICKS_PER_TEAM = 2
BULLPEN_STRONG_ERA = 3.4
BULLPEN_WEAK_ERA = 4.25
BULLPEN_STRONG_WHIP = 1.20
BULLPEN_WEAK_WHIP = 1.35
TEAM_K_LOW = 0.195
TEAM_K_HIGH = 0.235

def print_step(msg: str) -> None:
    print(msg, flush=True)

def get_json(url: str, params=None):
    r = requests.get(url, params=params, timeout=30)
    r.raise_for_status()
    return r.json()

def nz(x, fallback=0.0):
    return fallback if x is None or pd.isna(x) else x


def is_missing_value(x) -> bool:
    """True for None, pandas/NumPy NaN, blank strings, or string placeholders."""
    try:
        if x is None or pd.isna(x):
            return True
    except Exception:
        if x is None:
            return True
    s = str(x).strip().lower()
    return s in ("", "nan", "none", "null", "tbd")


def safe_int_value(x, default=None):
    """Convert MLB IDs/order fields safely without crashing on NaN/TBD blanks."""
    if is_missing_value(x):
        return default
    try:
        return int(float(x))
    except Exception:
        return default


def pct(h, ab):
    try:
        h = float(h)
        ab = float(ab)
        if ab <= 0:
            return None
        return round(100.0 * h / ab, 1)
    except Exception:
        return None

def normalize_name(v: str) -> str:
    if v is None:
        return ""
    s = str(v).strip().lower().replace("’", "'").replace("`", "'")
    return re.sub(r"[^a-z0-9]+", "", s)

def park_value(s: str) -> float:
    return {"Favorable": 10.0, "Neutral": 5.0, "Unfavorable": 0.0}.get(s or "Neutral", 5.0)

def overdue_value(status: str) -> float:
    if not status:
        return 0.0
    s = str(status).lower()
    if s.startswith("overdue"):
        m = re.search(r"\+(\d+)", status)
        return 8.0 + (float(m.group(1)) * 0.5 if m else 0.0)
    if s.startswith("slightly overdue"):
        m = re.search(r"\+(\d+)", status)
        return 5.0 + (float(m.group(1)) * 0.25 if m else 0.0)
    return 0.0

def innings_to_float(ip):
    if ip in (None, ""):
        return None
    if isinstance(ip, (int, float)):
        return float(ip)
    s = str(ip).strip()
    if "." not in s:
        try:
            return float(s)
        except Exception:
            return None
    whole, frac = s.split(".", 1)
    try:
        whole_i = int(whole)
        frac_i = int(frac)
    except Exception:
        try:
            return float(s)
        except Exception:
            return None
    return whole_i + {0: 0.0, 1: 1 / 3, 2: 2 / 3}.get(frac_i, 0.0)


def safe_div(n, d, fallback=0.0):
    try:
        n = float(n)
        d = float(d)
        if d == 0:
            return fallback
        return n / d
    except Exception:
        return fallback

def get_team_volatility(team_name: str) -> float:
    return float(TEAM_VOLATILITY_MAP.get(team_name, 1.0))

def get_public_bias(team_name: str) -> float:
    return float(PUBLIC_BIAS_MAP.get(team_name, 1.0))

def get_volatility_penalty(team_name: str, mode: str) -> float:
    vol = get_team_volatility(team_name)
    if vol <= 1.0:
        return 0.0
    if mode == "hit":
        return round((vol - 1.0) * 3.0, 3)
    if mode == "ml":
        return round((vol - 1.0) * 2.0, 3)
    return round((vol - 1.0) * 1.5, 3)

def get_public_bias_penalty(team_name: str, mode: str) -> float:
    pb = get_public_bias(team_name)
    if pb <= 1.0:
        return 0.0
    if mode == "ml":
        return round((pb - 1.0) * 2.0, 3)
    if mode == "hr":
        return round((pb - 1.0) * 1.25, 3)
    return round((pb - 1.0), 3)

def get_pitcher_game_logs(player_id: int, season: int) -> pd.DataFrame:
    data = get_json(
        f"https://statsapi.mlb.com/api/v1/people/{player_id}/stats",
        params={"stats": "gameLog", "group": "pitching", "season": season, "gameType": "R"},
    )
    stats = data.get("stats") or []
    if not stats:
        return pd.DataFrame(columns=["date", "inningsPitched", "strikeOuts", "pitchesThrown", "battersFaced", "earnedRuns", "hitsAllowed", "walks"])
    rows = []
    for s in stats[0].get("splits", []) or []:
        stat = s.get("stat", {})
        rows.append({
            "date": pd.to_datetime(s.get("date")).normalize() if s.get("date") else pd.NaT,
            "inningsPitched": innings_to_float(stat.get("inningsPitched")),
            "strikeOuts": float(stat.get("strikeOuts", 0) or 0),
            "pitchesThrown": float(stat.get("numberOfPitches", 0) or 0),
            "battersFaced": float(stat.get("battersFaced", 0) or 0),
            "earnedRuns": float(stat.get("earnedRuns", 0) or 0),
            "hitsAllowed": float(stat.get("hits", 0) or 0),
            "walks": float(stat.get("baseOnBalls", 0) or 0),
        })
    return pd.DataFrame(rows).sort_values("date").reset_index(drop=True)

def summarize_recent_pitcher_form(logs: pd.DataFrame) -> dict:
    if logs is None or logs.empty:
        return {
            "last2_ip_avg": None, "last3_ip_avg": None, "last2_k_avg": None, "last3_k_avg": None,
            "last2_pitch_avg": None, "last_start_ip": None, "last_start_k": None, "last_start_pitch_count": None,
            "last2_under5_count": None, "short_leash_flag": "Unknown", "recent_form_score": 0.0,
            "k5plus_last10_pct": None, "k6plus_last10_pct": None, "k7plus_last10_pct": None,
        }
    last2 = logs.tail(2).copy()
    last3 = logs.tail(3).copy()
    last2_ip_avg = round(last2["inningsPitched"].dropna().mean(), 3) if not last2.empty else None
    last3_ip_avg = round(last3["inningsPitched"].dropna().mean(), 3) if not last3.empty else None
    last2_k_avg = round(last2["strikeOuts"].dropna().mean(), 3) if not last2.empty else None
    last3_k_avg = round(last3["strikeOuts"].dropna().mean(), 3) if not last3.empty else None
    last2_pitch_avg = round(last2["pitchesThrown"].dropna().mean(), 3) if not last2.empty else None
    last_start = logs.tail(1).iloc[0]
    under5 = int((last2["inningsPitched"].fillna(0) < 5).sum()) if not last2.empty else None
    short_flag = "No"
    if len(last2) < 2:
        short_flag = "Unknown"
    elif under5 >= 2:
        short_flag = "Yes - last 2 starts under 5 IP"
    elif nz(last2_ip_avg) < 5:
        short_flag = "Yes - recent IP under 5"
    elif nz(last2_pitch_avg) and nz(last2_pitch_avg) < 85:
        short_flag = "Yes - pitch count risk"
    recent_form_score = round((nz(last2_ip_avg) * 1.4) + (nz(last2_k_avg) * 1.1) + (nz(last2_pitch_avg) * 0.03) - (under5 or 0) * 1.5, 3)
    k5plus_last10_pct = event_rate_last_n(logs, "strikeOuts", 10, 5)
    k6plus_last10_pct = event_rate_last_n(logs, "strikeOuts", 10, 6)
    k7plus_last10_pct = event_rate_last_n(logs, "strikeOuts", 10, 7)
    return {
        "last2_ip_avg": last2_ip_avg,
        "last3_ip_avg": last3_ip_avg,
        "last2_k_avg": last2_k_avg,
        "last3_k_avg": last3_k_avg,
        "last2_pitch_avg": last2_pitch_avg,
        "last_start_ip": last_start.get("inningsPitched"),
        "last_start_k": last_start.get("strikeOuts"),
        "last_start_pitch_count": last_start.get("pitchesThrown"),
        "last2_under5_count": under5,
        "short_leash_flag": short_flag,
        "recent_form_score": recent_form_score,
        "k5plus_last10_pct": k5plus_last10_pct,
        "k6plus_last10_pct": k6plus_last10_pct,
        "k7plus_last10_pct": k7plus_last10_pct,
    }

def apply_team_pick_caps(df: pd.DataFrame, max_per_team: int = MAX_REFINED_PICKS_PER_TEAM) -> pd.DataFrame:
    if df is None or df.empty or "teamName" not in df.columns:
        return df
    frames = []
    for _, grp in df.groupby("teamName", sort=False):
        frames.append(grp.head(max_per_team))
    if not frames:
        return df.iloc[0:0].copy()
    return pd.concat(frames, ignore_index=True)


def classify_team_k_tendency(k_rate: float) -> str:
    kr = nz(k_rate, None)
    if kr is None:
        return "Unknown"
    if kr >= TEAM_K_HIGH:
        return "High K"
    if kr <= TEAM_K_LOW:
        return "Low K"
    return "Neutral"


def classify_bullpen_grade(era: float, whip: float) -> str:
    e = nz(era, None)
    w = nz(whip, None)
    if e is None and w is None:
        return "Unknown"
    if (e is not None and e <= BULLPEN_STRONG_ERA) and (w is None or w <= BULLPEN_STRONG_WHIP):
        return "Strong"
    if (e is not None and e >= BULLPEN_WEAK_ERA) or (w is not None and w >= BULLPEN_WEAK_WHIP):
        return "Weak"
    return "Neutral"


def bullpen_hitter_adjustment(grade: str, mode: str = "hit") -> float:
    g = str(grade or "Unknown")
    if g == "Weak":
        return 0.35 if mode == "hit" else 0.20
    if g == "Strong":
        return -0.25 if mode == "hit" else -0.15
    return 0.0


def bullpen_pitcher_adjustment(grade: str) -> float:
    g = str(grade or "Unknown")
    if g == "Strong":
        return 0.35
    if g == "Weak":
        return -0.35
    return 0.0


def k_matchup_bonus_from_rate(k_rate: float) -> float:
    kr = nz(k_rate, None)
    if kr is None:
        return 0.0
    if kr >= 0.245:
        return 1.1
    if kr >= 0.235:
        return 0.7
    if kr >= 0.225:
        return 0.35
    if kr <= 0.185:
        return -0.7
    if kr <= 0.195:
        return -0.35
    return 0.0


def get_team_hitting_context(team_id: int, season: int) -> dict:
    try:
        data = get_json(
            "https://statsapi.mlb.com/api/v1/stats",
            params={
                "stats": "season",
                "group": "hitting",
                "season": season,
                "gameType": "R",
                "teamId": team_id,
            },
        )
        splits = data.get("stats", [{}])[0].get("splits", []) or []
        if not splits:
            return {}
        stat = splits[0].get("stat") or {}
        games = float(stat.get("gamesPlayed", 0) or 0)
        strikeouts = float(stat.get("strikeOuts", 0) or 0)
        at_bats = float(stat.get("atBats", 0) or 0)
        walks = float(stat.get("baseOnBalls", 0) or 0)
        hbp = float(stat.get("hitByPitch", 0) or 0)
        sac_flies = float(stat.get("sacFlies", 0) or 0)
        pa = at_bats + walks + hbp + sac_flies
        k_rate = round(safe_div(strikeouts, pa, None), 4) if pa else None
        return {
            "team_k_rate": k_rate,
            "team_k_per_game": round(safe_div(strikeouts, games, None), 3) if games else None,
            "team_pa": pa,
            "team_k_tendency": classify_team_k_tendency(k_rate),
        }
    except Exception:
        return {}


def _extract_pitching_stat_block(team_id: int, season: int, sit_codes: str | None = None) -> dict:
    params = {
        "stats": "season",
        "group": "pitching",
        "season": season,
        "gameType": "R",
        "teamId": team_id,
    }
    if sit_codes:
        params["sitCodes"] = sit_codes
    data = get_json("https://statsapi.mlb.com/api/v1/stats", params=params)
    stats = data.get("stats") or []
    splits = stats[0].get("splits", []) if stats else []
    if not splits:
        return {}
    return splits[0].get("stat") or {}


def get_team_pitching_context(team_id: int, season: int) -> dict:
    total_stat = {}
    relief_stat = {}
    try:
        total_stat = _extract_pitching_stat_block(team_id, season)
    except Exception:
        total_stat = {}
    try:
        relief_stat = _extract_pitching_stat_block(team_id, season, sit_codes="rp")
    except Exception:
        relief_stat = {}
    stat = relief_stat or total_stat
    if not stat:
        return {}
    bullpen_era = float(stat.get("era", 0) or 0) if stat.get("era") not in (None, "") else None
    bullpen_whip = float(stat.get("whip", 0) or 0) if stat.get("whip") not in (None, "") else None
    return {
        "bullpen_era": bullpen_era,
        "bullpen_whip": bullpen_whip,
        "bullpen_grade": classify_bullpen_grade(bullpen_era, bullpen_whip),
        "bullpen_source": "relief_split" if relief_stat else "team_total_fallback",
        "team_pitching_era": float(total_stat.get("era", 0) or 0) if total_stat.get("era") not in (None, "") else bullpen_era,
        "team_pitching_whip": float(total_stat.get("whip", 0) or 0) if total_stat.get("whip") not in (None, "") else bullpen_whip,
    }


def build_team_context_df(schedule_rows: pd.DataFrame, season: int) -> pd.DataFrame:
    team_map = get_team_map(schedule_rows)
    rows = []
    total = max(len(team_map), 1)
    for i, (team_name, team_id) in enumerate(team_map.items(), 1):
        print_step(f"📊 Team context {i}/{total}: {team_name}")
        hit_ctx = get_team_hitting_context(team_id, season)
        pitch_ctx = get_team_pitching_context(team_id, season)
        rows.append({"teamName": team_name, **hit_ctx, **pitch_ctx})
        time.sleep(SLEEP_BETWEEN_CALLS)
    return pd.DataFrame(rows)


def enrich_player_rows_with_team_context(player_rows: pd.DataFrame, pitcher_metrics: pd.DataFrame, team_context_df: pd.DataFrame) -> pd.DataFrame:
    if player_rows is None or player_rows.empty:
        return player_rows
    rows = player_rows.copy()
    if pitcher_metrics is not None and not pitcher_metrics.empty:
        opp_lookup = pitcher_metrics[["teamName", "opponentTeam"]].drop_duplicates().rename(columns={"teamName": "opponentTeam", "opponentTeam": "teamName"})
        rows = rows.merge(opp_lookup, on="teamName", how="left")
    if team_context_df is not None and not team_context_df.empty:
        offense_ctx = team_context_df[["teamName", "team_k_rate", "team_k_per_game", "team_k_tendency"]].drop_duplicates()
        opp_ctx = team_context_df[["teamName", "bullpen_era", "bullpen_whip", "bullpen_grade", "bullpen_source", "team_pitching_era", "team_pitching_whip"]].drop_duplicates().rename(columns={
            "teamName": "opponentTeam",
            "bullpen_era": "opp_bullpen_era",
            "bullpen_whip": "opp_bullpen_whip",
            "bullpen_grade": "opp_bullpen_grade",
            "bullpen_source": "opp_bullpen_source",
            "team_pitching_era": "opp_team_pitching_era",
            "team_pitching_whip": "opp_team_pitching_whip",
        })
        rows = rows.merge(offense_ctx, on="teamName", how="left")
        rows = rows.merge(opp_ctx, on="opponentTeam", how="left")
    rows["k_tendency_hit_penalty"] = rows["team_k_rate"].apply(lambda x: round(max(0.0, nz(x) - 0.215) * 8.0, 3) if pd.notna(x) else 0.0)
    rows["bullpen_hit_adjustment"] = rows["opp_bullpen_grade"].apply(lambda x: bullpen_hitter_adjustment(x, "hit"))
    rows["bullpen_hr_adjustment"] = rows["opp_bullpen_grade"].apply(lambda x: bullpen_hitter_adjustment(x, "hr"))
    rows["Hit_score"] = (rows["Hit_score"].fillna(0) - rows["k_tendency_hit_penalty"] + rows["bullpen_hit_adjustment"]).round(3)
    rows["HR_score"] = (rows["HR_score"].fillna(0) + rows["bullpen_hr_adjustment"]).round(3)
    return rows


def enrich_pitcher_metrics_with_team_context(pitcher_metrics: pd.DataFrame, team_context_df: pd.DataFrame) -> pd.DataFrame:
    if pitcher_metrics is None or pitcher_metrics.empty:
        return pitcher_metrics
    rows = pitcher_metrics.copy()
    if team_context_df is not None and not team_context_df.empty:
        opp_hit_ctx = team_context_df[["teamName", "team_k_rate", "team_k_tendency"]].drop_duplicates().rename(columns={
            "teamName": "opponentTeam",
            "team_k_rate": "opp_team_k_rate",
            "team_k_tendency": "opp_team_k_tendency",
        })
        own_pen_ctx = team_context_df[["teamName", "bullpen_era", "bullpen_whip", "bullpen_grade", "bullpen_source"]].drop_duplicates().rename(columns={
            "bullpen_era": "own_bullpen_era",
            "bullpen_whip": "own_bullpen_whip",
            "bullpen_grade": "own_bullpen_grade",
            "bullpen_source": "own_bullpen_source",
        })
        rows = rows.merge(opp_hit_ctx, on="opponentTeam", how="left")
        rows = rows.merge(own_pen_ctx, on="teamName", how="left")
    rows["opp_k_matchup_bonus"] = rows["opp_team_k_rate"].apply(k_matchup_bonus_from_rate)
    rows["bullpen_support_adjustment"] = rows["own_bullpen_grade"].apply(bullpen_pitcher_adjustment)
    rows["pitcher_score_adj"] = (rows["pitcher_score_adj"].fillna(0) + rows["opp_k_matchup_bonus"] + rows["bullpen_support_adjustment"]).round(3)
    def _upgrade_pick_type(row):
        current = str(row.get("pick_type") or "Neutral")
        if str(row.get("short_leash_flag") or "").startswith("Yes"):
            return "Short Leash Risk"
        bonus = nz(row.get("opp_k_matchup_bonus"))
        score = nz(row.get("pitcher_score_adj"))
        if current == "Neutral" and bonus >= 0.7 and score >= 5.5:
            return "K Upside"
        if current == "K Upside" and bonus >= 0.7 and score >= 6.5:
            return "Strong SP"
        if current == "Strong SP" and bonus <= -0.35:
            return "Neutral"
        return current
    rows["pick_type"] = rows.apply(_upgrade_pick_type, axis=1)
    return rows

def get_schedule_rows(target_date: str) -> pd.DataFrame:
    print_step(f"📅 Pulling schedule for {target_date} ...")
    data = get_json(
        "https://statsapi.mlb.com/api/v1/schedule",
        params={"sportId": 1, "date": target_date, "hydrate": "team,probablePitcher,venue"},
    )
    rows = []
    for d in data.get("dates", []):
        for g in d.get("games", []):
            teams = g.get("teams", {})
            home = teams.get("home", {})
            away = teams.get("away", {})
            game_dt_raw = g.get("gameDate")
            game_time_et = None
            if game_dt_raw:
                try:
                    game_dt_et = pd.to_datetime(game_dt_raw, utc=True).tz_convert("America/New_York")
                    game_time_et = game_dt_et.strftime("%-I:%M %p ET")
                except Exception:
                    try:
                        game_time_et = pd.to_datetime(game_dt_raw, utc=True).tz_convert("America/New_York").strftime("%I:%M %p ET").lstrip("0")
                    except Exception:
                        game_time_et = None
            rows.append({
                "game_date": target_date,
                "game_datetime_utc": game_dt_raw,
                "game_time_et": game_time_et,
                "away_team": (away.get("team") or {}).get("name"),
                "home_team": (home.get("team") or {}).get("name"),
                "venue": (g.get("venue") or {}).get("name"),
                "away_probable_pitcher": (away.get("probablePitcher") or {}).get("fullName"),
                "away_probable_pitcher_id": (away.get("probablePitcher") or {}).get("id"),
                "home_probable_pitcher": (home.get("probablePitcher") or {}).get("fullName"),
                "home_probable_pitcher_id": (home.get("probablePitcher") or {}).get("id"),
                "gamePk": g.get("gamePk"),
            })
    return pd.DataFrame(rows)


def filter_pregame_schedule_rows(schedule_rows: pd.DataFrame, now_et=None, buffer_minutes: int = 0) -> pd.DataFrame:
    """Keep only games that have not started yet for actionable Final Card picks."""
    if schedule_rows is None or schedule_rows.empty:
        return schedule_rows.copy() if schedule_rows is not None else pd.DataFrame()

    rows = schedule_rows.copy()
    if now_et is None:
        now_et = dt.datetime.now(ZoneInfo("America/New_York"))
    if now_et.tzinfo is None:
        now_et = now_et.replace(tzinfo=ZoneInfo("America/New_York"))

    try:
        slate_date = pd.to_datetime(rows["game_date"].dropna().iloc[0]).date()
        today_et = now_et.date()
        if slate_date > today_et:
            rows["game_lock_status"] = "Pregame - future date"
            return rows
        if slate_date < today_et:
            rows["game_lock_status"] = "Locked - past date"
            return rows.iloc[0:0].copy()
    except Exception:
        pass

    if "game_datetime_utc" not in rows.columns:
        rows["game_lock_status"] = "Locked - missing start time"
        return rows.iloc[0:0].copy()

    game_dt_utc = pd.to_datetime(rows["game_datetime_utc"], errors="coerce", utc=True)
    game_dt_et = game_dt_utc.dt.tz_convert("America/New_York")
    cutoff = now_et + dt.timedelta(minutes=buffer_minutes)
    mask = game_dt_et > cutoff

    rows["game_start_et_dt"] = game_dt_et
    rows["game_lock_status"] = mask.map(lambda x: "Pregame" if x else "Locked - already started")
    return rows[mask].copy()

def get_pitcher_hand(pid):
    safe_pid = safe_int_value(pid)
    if safe_pid is None:
        return None
    try:
        p = get_json(f"https://statsapi.mlb.com/api/v1/people/{safe_pid}")
        ppl = p.get("people", [])
        if ppl:
            ph = ppl[0].get("pitchHand") or {}
            code = (ph.get("code") or ph.get("description") or "").upper()
            if code.startswith("R"):
                return "R"
            if code.startswith("L"):
                return "L"
    except Exception:
        return None
    return None

def get_schedule_game_context(target_date: str):
    schedule_rows = get_schedule_rows(target_date)
    ctx = {}
    for _, g in schedule_rows.iterrows():
        home_team = g.get("home_team")
        away_team = g.get("away_team")
        pf = PARK_FAVORABILITY_MAP.get(home_team, "Neutral")
        ctx[home_team] = {
            "opp_pitcher_name": g.get("away_probable_pitcher"),
            "opp_pitcher_id": g.get("away_probable_pitcher_id"),
            "opp_pitcher_hand": get_pitcher_hand(g.get("away_probable_pitcher_id")),
            "game_park_team": home_team,
            "game_park_name": g.get("venue"),
            "park_favorability": pf,
        }
        ctx[away_team] = {
            "opp_pitcher_name": g.get("home_probable_pitcher"),
            "opp_pitcher_id": g.get("home_probable_pitcher_id"),
            "opp_pitcher_hand": get_pitcher_hand(g.get("home_probable_pitcher_id")),
            "game_park_team": home_team,
            "game_park_name": g.get("venue"),
            "park_favorability": pf,
        }
    return ctx, schedule_rows

def get_team_roster(team_id: int, season: int):
    data = get_json(
        f"https://statsapi.mlb.com/api/v1/teams/{team_id}/roster",
        params={"rosterType": "fullSeason", "season": season},
    )
    roster = data.get("roster", []) or []
    out = {}
    for r in roster:
        person = r.get("person") or {}
        pos = (r.get("position") or {}).get("abbreviation", "")
        out[int(person.get("id"))] = {"playerName": person.get("fullName"), "pos": pos}
    return out

def get_team_map(schedule_rows: pd.DataFrame):
    data = get_json("https://statsapi.mlb.com/api/v1/teams", params={"sportId": 1})
    teams = data.get("teams", []) or []
    team_lookup = {t.get("name"): int(t.get("id")) for t in teams if t.get("name") and t.get("id")}
    scheduled = set(schedule_rows["home_team"].dropna().tolist() + schedule_rows["away_team"].dropna().tolist())
    return {name: team_lookup[name] for name in scheduled if name in team_lookup}

def get_team_hitting_pool(team_id: int, season: int):
    data = get_json(
        "https://statsapi.mlb.com/api/v1/stats",
        params={
            "stats": "season",
            "group": "hitting",
            "season": season,
            "gameType": "R",
            "teamId": team_id,
        },
    )
    splits = data.get("stats", [{}])[0].get("splits", []) or []
    rows = []
    for s in splits:
        player = s.get("player", {})
        stat = s.get("stat", {})
        rows.append({
            "playerId": int(player.get("id")),
            "playerName": player.get("fullName"),
            "homeRuns": int(stat.get("homeRuns", 0)),
            "hits": int(stat.get("hits", 0)),
            "gamesPlayed": int(stat.get("gamesPlayed", 0)),
            "atBats": int(stat.get("atBats", 0)),
        })
    return pd.DataFrame(rows)

def build_scheduled_player_pool(schedule_rows: pd.DataFrame, season: int):
    print_step("📡 Pulling all scheduled-team hitters ...")
    team_map = get_team_map(schedule_rows)
    frames = []
    for i, (team_name, team_id) in enumerate(team_map.items(), 1):
        print_step(f"🏟️ Team {i}/{len(team_map)}: {team_name}")
        pool = get_team_hitting_pool(team_id, season)
        if not pool.empty:
            pool["teamName"] = team_name
            frames.append(pool)
        time.sleep(SLEEP_BETWEEN_CALLS)
    if not frames:
        return pd.DataFrame(columns=["playerId", "playerName", "homeRuns", "hits", "gamesPlayed", "atBats", "teamName"])
    df = pd.concat(frames, ignore_index=True)
    return df.drop_duplicates(subset=["playerId", "teamName"]).reset_index(drop=True)

def get_player_game_logs(player_id: int, season: int) -> pd.DataFrame:
    data = get_json(
        f"https://statsapi.mlb.com/api/v1/people/{player_id}/stats",
        params={"stats": "gameLog", "group": "hitting", "season": season, "gameType": "R"},
    )
    stats = data.get("stats") or []
    if not stats:
        return pd.DataFrame(columns=["date", "homeRuns", "hits"])
    rows = []
    for s in stats[0].get("splits", []) or []:
        stat = s.get("stat", {})
        rows.append({
            "date": pd.to_datetime(s.get("date")).normalize() if s.get("date") else pd.NaT,
            "homeRuns": int(stat.get("homeRuns", 0) or 0),
            "hits": int(stat.get("hits", 0) or 0),
            "totalBases": int(stat.get("totalBases", 0) or 0),
            "rbi": int(stat.get("rbi", 0) or 0),
            "runs": int(stat.get("runs", 0) or 0),
            "atBats": int(stat.get("atBats", 0) or 0),
            "doubles": int(stat.get("doubles", 0) or 0),
            "triples": int(stat.get("triples", 0) or 0),
            "baseOnBalls": int(stat.get("baseOnBalls", 0) or 0),
            "strikeOuts": int(stat.get("strikeOuts", 0) or 0),
        })
    return pd.DataFrame(rows).sort_values("date").reset_index(drop=True)

def compute_drought_metrics(df: pd.DataFrame, col: str) -> dict:
    total = len(df)
    if total == 0:
        return {"last_event_date": None, "current_gap": None, "avg_games_between": None, "longest_drought": None}
    idxs = df.index[df[col] > 0].tolist()
    if not idxs:
        return {"last_event_date": None, "current_gap": total, "avg_games_between": None, "longest_drought": total}
    last_idx = idxs[-1]
    current_gap = total - last_idx - 1
    last_event_date = df.loc[last_idx, "date"]
    if len(idxs) == 1:
        longest = max(idxs[0], current_gap)
        return {"last_event_date": last_event_date.date() if pd.notna(last_event_date) else None, "current_gap": current_gap, "avg_games_between": None, "longest_drought": longest}
    gaps = [idxs[i + 1] - idxs[i] - 1 for i in range(len(idxs) - 1)]
    return {"last_event_date": last_event_date.date() if pd.notna(last_event_date) else None, "current_gap": current_gap, "avg_games_between": round(sum(gaps) / len(gaps), 2) if gaps else None, "longest_drought": max(idxs[0], current_gap, max(gaps))}



def event_rate_last_n(logs: pd.DataFrame, col: str, n: int, threshold: float = 1) -> float | None:
    """Percent of last N games where col >= threshold."""
    if logs is None or logs.empty or col not in logs.columns:
        return None
    tail = logs.tail(n)
    if len(tail) < n:
        return None
    return round(float((pd.to_numeric(tail[col], errors="coerce").fillna(0) >= threshold).mean() * 100), 1)


def current_event_streak(logs: pd.DataFrame, col: str, threshold: float = 1) -> int:
    """Current consecutive games from most recent backwards where col >= threshold."""
    if logs is None or logs.empty or col not in logs.columns:
        return 0
    streak = 0
    vals = pd.to_numeric(logs[col], errors="coerce").fillna(0).tolist()[::-1]
    for v in vals:
        if v >= threshold:
            streak += 1
        else:
            break
    return streak


def contact_momentum_bonus(hit_l5, hit_l10, streak) -> float:
    bonus = 0.0
    if hit_l10 is not None:
        if hit_l10 >= 90: bonus += 0.45
        elif hit_l10 >= 80: bonus += 0.30
        elif hit_l10 >= 70: bonus += 0.15
        elif hit_l10 >= 60: bonus += 0.05
    if hit_l5 is not None:
        if hit_l5 >= 100: bonus += 0.30
        elif hit_l5 >= 80: bonus += 0.20
        elif hit_l5 <= 40: bonus -= 0.20
    if streak >= 5: bonus += 0.15
    elif streak >= 3: bonus += 0.08
    return round(bonus, 3)


def add_contact_quality_engine(player_rows: pd.DataFrame) -> pd.DataFrame:
    """Add a process/form layer so picks are not only outcome/drought based.

    This is a practical contact-quality proxy using data already available from MLB game logs:
    - last 10 hit rate
    - last 5 hit rate
    - 2+ total-base rate
    - current hit streak
    - lineup slot
    - opposing bullpen grade
    - opponent pitcher label

    It does not require Statcast, but it moves the model closer to a true
    process-based projection instead of a coin-flip recent-result model.
    """
    if player_rows is None or player_rows.empty:
        return player_rows

    df = player_rows.copy()

    def num(col, default=0.0):
        if col not in df.columns:
            return pd.Series([default] * len(df), index=df.index)
        return pd.to_numeric(df[col], errors="coerce").fillna(default)

    hit10 = num("hit_pct_last_10")
    hit5 = num("hit_pct_last_5")
    tb10 = num("tb2plus_last10_pct")
    streak = num("current_hit_streak")
    slot = num("batting_order_slot", 9)
    hit_score = num("Hit_score")
    hr_score = num("HR_score")

    opp_type = df["opponent_pitcher_pick_type"].astype(str) if "opponent_pitcher_pick_type" in df.columns else pd.Series(["Neutral"] * len(df), index=df.index)
    bullpen = df["opp_bullpen_grade"].astype(str) if "opp_bullpen_grade" in df.columns else pd.Series(["Unknown"] * len(df), index=df.index)
    lineup_status = df["lineup_status"].astype(str) if "lineup_status" in df.columns else pd.Series(["Unknown"] * len(df), index=df.index)

    contact_score = (
        (hit10 / 100.0) * 2.20
        + (hit5 / 100.0) * 1.05
        + (tb10 / 100.0) * 0.80
        + streak.clip(upper=6) * 0.10
        + (10 - slot.clip(lower=1, upper=9)) * 0.08
    )

    contact_score += bullpen.map({"Weak": 0.35, "Neutral": 0.05, "Strong": -0.30, "Unknown": 0.0}).fillna(0.0)
    contact_score += opp_type.map({
        "Short Leash Risk": 0.25,
        "Attack With Hitters": 0.35,
        "Low Sample": 0.15,
        "Neutral": 0.0,
        "K Upside": -0.20,
        "Strong SP": -0.75,
    }).fillna(0.0)

    # Penalize cold/volatile profiles so they do not feel random.
    cold_penalty = ((hit10 < 60).astype(int) * 0.45) + ((hit5 <= 40).astype(int) * 0.35) + ((streak == 0).astype(int) * 0.15)
    contact_score = (contact_score - cold_penalty).round(3)

    df["contact_quality_score"] = contact_score
    df["hit_quality_label"] = pd.cut(
        contact_score,
        bins=[-999, 1.35, 1.85, 2.35, 999],
        labels=["Weak", "Playable", "Strong", "Elite"]
    ).astype(str)

    df["hit_quality_gate"] = (
        lineup_status.eq("Confirmed Starter")
        & (slot <= 6)
        & (hit_score >= 3.75)
        & (hit10 >= 60)
        & (~opp_type.eq("Strong SP"))
        & (contact_score >= 1.55)
    )

    # HR contact proxy: HRs should need power/contact support, not drought alone.
    df["hr_contact_proxy"] = (
        (tb10 / 100.0) * 1.60
        + (hr_score / 5.0).clip(lower=0, upper=2.0) * 0.70
        + (hit10 / 100.0) * 0.35
        + bullpen.map({"Weak": 0.30, "Neutral": 0.05, "Strong": -0.20, "Unknown": 0.0}).fillna(0.0)
        + opp_type.map({"Short Leash Risk": 0.20, "Attack With Hitters": 0.25, "Low Sample": 0.15, "Strong SP": -0.50}).fillna(0.0)
    ).round(3)

    # Apply small score adjustments after the proxy fields are created.
    df["Hit_score"] = (hit_score + (df["contact_quality_score"] - 1.75).clip(lower=-0.60, upper=0.75)).round(3)
    df["HR_score"] = (hr_score + (df["hr_contact_proxy"] - 1.25).clip(lower=-0.45, upper=0.55)).round(3)

    return df


def determine_status(current_gap, avg_gap):
    if avg_gap is None or current_gap is None:
        return "N/A"
    if current_gap <= avg_gap:
        return "On Pace"
    if current_gap <= 1.5 * avg_gap:
        return f"Slightly Overdue (+{current_gap - int(avg_gap)})"
    return f"Overdue (+{current_gap - int(avg_gap)})"

def average_games_per_event(games_played, event_count):
    gp = nz(games_played, None)
    ec = nz(event_count, None)
    if gp is None or ec is None or ec <= 0:
        return None
    return round(float(gp) / float(ec), 2)


def get_pitcher_season_stats(pid: int, season: int) -> dict:
    safe_pid = safe_int_value(pid)
    if safe_pid is None:
        return {}
    try:
        data = get_json(
            f"https://statsapi.mlb.com/api/v1/people/{safe_pid}/stats",
            params={"stats": "season", "group": "pitching", "season": season, "gameType": "R"},
        )
        stats = data.get("stats") or []
        splits = stats[0].get("splits", []) if stats else []
        if not splits:
            return {}
        stat = splits[0].get("stat") or {}
        return {
            "inningsPitched": innings_to_float(stat.get("inningsPitched")),
            "strikeOuts": float(stat.get("strikeOuts", 0) or 0),
            "earnedRuns": float(stat.get("earnedRuns", 0) or 0),
            "hitsAllowed": float(stat.get("hits", 0) or 0),
            "baseOnBalls": float(stat.get("baseOnBalls", 0) or 0),
            "gamesStarted": float(stat.get("gamesStarted", 0) or 0),
            "era": float(stat.get("era", 0) or 0) if stat.get("era") not in (None, "") else None,
            "whip": float(stat.get("whip", 0) or 0) if stat.get("whip") not in (None, "") else None,
        }
    except Exception:
        return {}

def compute_pitcher_score(ip, so, er, ha, bb):
    return round((nz(so) * 1.5) + (nz(ip) * 1.2) - (nz(er) * 2.0) - (nz(ha) * 0.8) - (nz(bb) * 1.2), 3)

def classify_pitcher_pick(score_adj, ip, so, er, ha):
    if nz(ip) < 3:
        return "Low Sample"
    if nz(score_adj) >= 6:
        return "Strong SP"
    if nz(so) >= 6 and nz(ip) >= 4:
        return "K Upside"
    if nz(er) >= 3 or nz(ha) >= 6:
        return "Attack With Hitters"
    return "Neutral"

def build_pitcher_metrics(schedule_rows: pd.DataFrame, season: int) -> pd.DataFrame:
    rows = []
    total = max(len(schedule_rows) * 2, 1)
    counter = 0
    for _, g in schedule_rows.iterrows():
        for team, opp, pitcher_name, pitcher_id in [
            (g.get("away_team"), g.get("home_team"), g.get("away_probable_pitcher"), g.get("away_probable_pitcher_id")),
            (g.get("home_team"), g.get("away_team"), g.get("home_probable_pitcher"), g.get("home_probable_pitcher_id")),
        ]:
            counter += 1
            safe_pid = safe_int_value(pitcher_id)
            if is_missing_value(team) or is_missing_value(opp):
                continue
            if is_missing_value(pitcher_name) or safe_pid is None:
                print_step(f"⚠️ Skipping TBD/missing probable pitcher for {team} vs {opp}")
                continue
            print_step(f"🎯 Pitcher {counter}/{total}: {pitcher_name} ({team})")
            stat = get_pitcher_season_stats(safe_pid, season)
            logs = get_pitcher_game_logs(safe_pid, season)
            recent = summarize_recent_pitcher_form(logs)
            ip = stat.get("inningsPitched")
            so = stat.get("strikeOuts")
            er = stat.get("earnedRuns")
            ha = stat.get("hitsAllowed")
            bb = stat.get("baseOnBalls")
            raw_score = compute_pitcher_score(ip, so, er, ha, bb)
            short_leash_penalty = 0.0
            if str(recent.get("short_leash_flag") or "").startswith("Yes"):
                short_leash_penalty = 3.0
            elif str(recent.get("short_leash_flag") or "") == "Unknown":
                short_leash_penalty = 1.0
            score_adj = round(raw_score - short_leash_penalty + nz(recent.get("recent_form_score")) * 0.15, 3)
            pick_type = classify_pitcher_pick(score_adj, ip, so, er, ha)
            if str(recent.get("short_leash_flag") or "").startswith("Yes"):
                pick_type = "Short Leash Risk"
            rows.append({
                "pitcherName": pitcher_name, "teamName": team, "opponentTeam": opp,
                "innings_pitched": ip, "strikeouts": so, "earned_runs": er, "hits_allowed": ha, "walks": bb,
                "games_started": stat.get("gamesStarted"), "era": stat.get("era"), "whip": stat.get("whip"),
                "pitcher_score": raw_score, "pitcher_score_adj": score_adj,
                "recent_form_score": recent.get("recent_form_score"),
                "last2_ip_avg": recent.get("last2_ip_avg"), "last3_ip_avg": recent.get("last3_ip_avg"),
                "last2_k_avg": recent.get("last2_k_avg"), "last3_k_avg": recent.get("last3_k_avg"),
                "last2_pitch_avg": recent.get("last2_pitch_avg"),
                "last_start_ip": recent.get("last_start_ip"), "last_start_k": recent.get("last_start_k"),
                "last_start_pitch_count": recent.get("last_start_pitch_count"),
                "last2_under5_count": recent.get("last2_under5_count"),
                "k5plus_last10_pct": recent.get("k5plus_last10_pct"),
                "k6plus_last10_pct": recent.get("k6plus_last10_pct"),
                "k7plus_last10_pct": recent.get("k7plus_last10_pct"),
                "short_leash_flag": recent.get("short_leash_flag"),
                "sample_flag": "Low Sample" if nz(ip) < 3 else "OK",
                "pick_type": pick_type, "probable_starter_name": pitcher_name, "starter_status": "Confirmed",
            })
            time.sleep(SLEEP_BETWEEN_CALLS)
    cols = ["pitcherName","teamName","opponentTeam","innings_pitched","strikeouts","earned_runs","hits_allowed","walks","games_started","era","whip","pitcher_score","pitcher_score_adj","recent_form_score","last2_ip_avg","last3_ip_avg","last2_k_avg","last3_k_avg","last2_pitch_avg","last_start_ip","last_start_k","last_start_pitch_count","last2_under5_count","k5plus_last10_pct","k6plus_last10_pct","k7plus_last10_pct","short_leash_flag","sample_flag","pick_type","probable_starter_name","starter_status"]
    return pd.DataFrame(rows, columns=cols).sort_values(["pitcher_score_adj","pitcher_score"], ascending=False).reset_index(drop=True)

def get_confirmed_lineups(target_date: str):
    print_step("🧾 Pulling confirmed lineups ...")
    status_map = {}
    slot_map = {}
    try:
        schedule = get_json("https://statsapi.mlb.com/api/v1/schedule", params={"sportId": 1, "date": target_date})
        for d in schedule.get("dates", []):
            for g in d.get("games", []):
                game_pk = g.get("gamePk")
                if not game_pk:
                    continue
                try:
                    box = get_json(f"https://statsapi.mlb.com/api/v1/game/{game_pk}/boxscore")
                except Exception:
                    continue
                for side in ("home", "away"):
                    tblock = ((box.get("teams") or {}).get(side) or {})
                    team_name = ((tblock.get("team") or {}).get("name"))
                    batting_order = tblock.get("battingOrder") or []
                    players = tblock.get("players") or {}
                    if len(batting_order) < 9:
                        continue
                    for pid in batting_order:
                        p = players.get(f"ID{pid}") or {}
                        full_name = ((p.get("person") or {}).get("fullName"))
                        bo = str(p.get("battingOrder") or "").strip()
                        if team_name and full_name and bo:
                            key = (team_name, normalize_name(full_name))
                            status_map[key] = "Confirmed Starter"
                            try:
                                slot_map[key] = int(bo[:3]) // 100
                            except Exception:
                                slot_map[key] = None
    except Exception:
        pass
    return status_map, slot_map

def build_locked_player_pool(all_players: pd.DataFrame, lineup_map: dict, slot_map: dict) -> pd.DataFrame:
    print_step("🔒 Locking players to confirmed lineups ...")
    if all_players.empty:
        return all_players.copy()
    out = all_players.copy()
    out["lineup_status"] = out.apply(lambda r: lineup_map.get((r["teamName"], normalize_name(r["playerName"])), "Unknown"), axis=1)
    out["batting_order_slot"] = out.apply(lambda r: slot_map.get((r["teamName"], normalize_name(r["playerName"])), None), axis=1)
    out["starter_only_flag"] = out["lineup_status"].eq("Confirmed Starter")
    out = out[out["starter_only_flag"] == True].copy()
    out = out[out["batting_order_slot"].notna()].copy()
    return out

def build_hit_hr_rows(pool_df: pd.DataFrame, season: int, sched_ctx: dict) -> pd.DataFrame:
    rows = []
    total = max(len(pool_df), 1)
    for i, (_, row) in enumerate(pool_df.iterrows(), 1):
        print_step(f"👤 Player {i}/{total}: {row['playerName']} ({row['teamName']})")
        logs = get_player_game_logs(int(row["playerId"]), season)
        hr_d = compute_drought_metrics(logs, "homeRuns")
        hit_d = compute_drought_metrics(logs, "hits")
        avg_games_between_hrs = average_games_per_event(row.get("gamesPlayed"), row.get("homeRuns"))
        avg_games_between_hits = average_games_per_event(row.get("gamesPlayed"), row.get("hits"))
        hr_status = determine_status(hr_d["current_gap"], avg_games_between_hrs)
        hit_status = determine_status(hit_d["current_gap"], avg_games_between_hits)
        last10 = logs.tail(10)
        hit_pct_last_10 = event_rate_last_n(logs, "hits", 10, 1)
        hit_pct_last_5 = event_rate_last_n(logs, "hits", 5, 1)
        current_hit_streak = current_event_streak(logs, "hits", 1)
        tb2plus_last10_pct = event_rate_last_n(logs, "totalBases", 10, 2)
        tb2plus_last5_pct = event_rate_last_n(logs, "totalBases", 5, 2)
        run_last10_pct = event_rate_last_n(logs, "runs", 10, 1)
        rbi_last10_pct = event_rate_last_n(logs, "rbi", 10, 1)
        contact_momentum = contact_momentum_bonus(hit_pct_last_5, hit_pct_last_10, current_hit_streak)
        ctx = sched_ctx.get(row["teamName"], {})
        season_hit_pct = pct(row["hits"], row.get("atBats", 0))
        slot_raw = row.get("batting_order_slot")
        slot = 9 if pd.isna(slot_raw) else int(slot_raw)
        lineup_bonus = max(0, 10 - slot) * 0.12
        team_volatility = get_team_volatility(row["teamName"])
        public_bias = get_public_bias(row["teamName"])
        hit_vol_penalty = get_volatility_penalty(row["teamName"], "hit")
        hr_vol_penalty = get_volatility_penalty(row["teamName"], "hr")
        hr_public_penalty = get_public_bias_penalty(row["teamName"], "hr")
        hr_score_raw = (row["homeRuns"] / max(row["gamesPlayed"], 1) * 10 * 0.40) + (overdue_value(hr_status) * 0.25) + (park_value(ctx.get("park_favorability")) * 0.20) + lineup_bonus
        hit_score_raw = (nz(season_hit_pct) / 10.0 * 0.40) + (nz(hit_pct_last_10) / 10.0 * 0.20) + (park_value(ctx.get("park_favorability")) * 0.05) + lineup_bonus + contact_momentum
        hr_score = round(hr_score_raw - hr_vol_penalty - hr_public_penalty, 3)
        hit_score = round(hit_score_raw - hit_vol_penalty, 3)
        rows.append({
            "season": season, "teamName": row["teamName"], "playerName": row["playerName"], "playerId": row["playerId"],
            "homeRuns": row["homeRuns"], "gamesPlayed": row["gamesPlayed"], "totalHits": row["hits"],
            "avg_games_between_hrs": avg_games_between_hrs, "current_games_without_hr": hr_d["current_gap"],
            "longest_games_without_hr": hr_d["longest_drought"], "last_hr_date": hr_d["last_event_date"],
            "hr_status": hr_status, "avg_games_between_hits": avg_games_between_hits,
            "current_games_without_hit": hit_d["current_gap"], "longestHitDrought": hit_d["longest_drought"],
            "hit_status": hit_status, "hit_pct_last_10": hit_pct_last_10, "hit_pct_last_5": hit_pct_last_5,
            "current_hit_streak": current_hit_streak, "contact_momentum_bonus": contact_momentum,
            "tb2plus_last10_pct": tb2plus_last10_pct, "tb2plus_last5_pct": tb2plus_last5_pct,
            "run_last10_pct": run_last10_pct, "rbi_last10_pct": rbi_last10_pct,
            "season_hit_pct": season_hit_pct,
            "auto_pitcher_name": ctx.get("opp_pitcher_name"), "auto_pitcher_hand": ctx.get("opp_pitcher_hand"),
            "park_favorability": ctx.get("park_favorability"), "game_park_team": ctx.get("game_park_team"),
            "game_park_name": ctx.get("game_park_name"),
            "HR_score_raw": round(hr_score_raw, 3), "Hit_score_raw": round(hit_score_raw, 3),
            "HR_score": hr_score, "Hit_score": hit_score,
            "team_volatility": team_volatility, "public_bias": public_bias,
            "volatility_penalty_hit": hit_vol_penalty, "volatility_penalty_hr": hr_vol_penalty, "public_bias_penalty_hr": hr_public_penalty,
            "lineup_status": row.get("lineup_status"), "batting_order_slot": row.get("batting_order_slot"),
            "starter_only_flag": row.get("starter_only_flag"),
        })
        time.sleep(SLEEP_BETWEEN_CALLS)
    return pd.DataFrame(rows)

def build_game_rankings(schedule_rows, hr_rows, hit_rows, pitcher_metrics):
    hr_map = hr_rows.groupby("teamName")["HR_score"].mean().to_dict() if not hr_rows.empty else {}
    hit_map = hit_rows.groupby("teamName")["Hit_score"].mean().to_dict() if not hit_rows.empty else {}
    pmap = {r["teamName"]: r for _, r in pitcher_metrics.iterrows()} if not pitcher_metrics.empty else {}
    rows = []
    for _, g in schedule_rows.iterrows():
        for team, opp in [(g.get("away_team"), g.get("home_team")), (g.get("home_team"), g.get("away_team"))]:
            offense_hr = round(nz(hr_map.get(team)), 3)
            offense_hit = round(nz(hit_map.get(team)), 3)
            offense_score = round(offense_hr * 0.45 + offense_hit * 0.55, 3)
            p_self = pmap.get(team, {})
            p_opp = pmap.get(opp, {})
            vol_penalty_ml = get_volatility_penalty(team, "ml")
            public_penalty_ml = get_public_bias_penalty(team, "ml")
            short_leash_adj = 0.0
            if str(p_self.get("short_leash_flag") or "").startswith("Yes"):
                short_leash_adj = -2.0
            elif str(p_self.get("short_leash_flag") or "") == "Unknown":
                short_leash_adj = -0.5
            team_score = round((offense_score * 0.55) + (nz(p_self.get("pitcher_score_adj")) * 0.45) - vol_penalty_ml - public_penalty_ml + short_leash_adj, 3)
            rows.append({
                "game": f"{g.get('away_team')} @ {g.get('home_team')}",
                "game_time_et": g.get("game_time_et"),
                "game_datetime_utc": g.get("game_datetime_utc"),
                "teamName": team, "opponentTeam": opp, "venue": g.get("venue"),
                "offense_hr_score": offense_hr, "offense_hit_score": offense_hit, "offense_score": offense_score,
                "team_volatility": get_team_volatility(team), "public_bias": get_public_bias(team),
                "volatility_penalty_ml": vol_penalty_ml, "public_penalty_ml": public_penalty_ml,
                "pitcherName": p_self.get("pitcherName"), "pitcher_score": p_self.get("pitcher_score"),
                "pitcher_score_adj": p_self.get("pitcher_score_adj"), "pitcher_pick_type": p_self.get("pick_type"),
                "short_leash_flag": p_self.get("short_leash_flag"),
                "opponent_pitcher": p_opp.get("pitcherName"), "opponent_pitcher_score": p_opp.get("pitcher_score"),
                "opponent_pitcher_score_adj": p_opp.get("pitcher_score_adj"), "opponent_pitcher_pick_type": p_opp.get("pick_type"),
                "team_score": team_score,
            })
    df = pd.DataFrame(rows)
    if df.empty:
        return df
    gm = df.groupby("game")["team_score"].transform("mean")
    df["edge_vs_opponent"] = (df["team_score"] - (gm * 2 - df["team_score"])).round(3)
    def classify(row):
        edge = nz(row.get("edge_vs_opponent"))
        offense = nz(row.get("offense_score"))
        opp_pt = str(row.get("opponent_pitcher_pick_type") or "")
        own_pt = str(row.get("pitcher_pick_type") or "")
        own_short = str(row.get("short_leash_flag") or "")
        rating = "Strong" if edge >= 3 else "Lean" if edge >= 2.5 else "Fade" if edge <= -3 else "Slight Fade" if edge <= -2.5 else "Neutral"
        if own_short.startswith("Yes"):
            play = "Avoid"
        elif edge >= 2.5 and opp_pt == "Attack With Hitters":
            play = "Stack Spot"
        elif edge >= 2.5 and own_pt in ("Strong SP", "K Upside") and offense >= 2.2:
            play = "Moneyline Lean"
        elif edge <= -2.5:
            play = "Avoid"
        else:
            play = "Pass / Small Edge"
        return pd.Series([rating, play])
    df[["win_rating", "recommended_play"]] = df.apply(classify, axis=1)
    return df.sort_values(["game", "team_score"], ascending=[True, False]).reset_index(drop=True)

def build_refined_picks(player_rows, pitcher_metrics, game_rankings):
    """
    Refined Picks v2.2 — Rolling slate card, max 2 per game / max 10 total.

    Changes:
    - HARD cap: max 10 refined picks total.
    - Max 2 hitters per game so late games are not crowded out by early games.
    - Max 2 hitters per team.
    - Removes HR picks from Refined Picks; HR stays in Top Picks/Tier system only.
    - Requires valid game mapping.
    - Blocks Strong SP matchups.
    - If lineups are confirmed, prioritizes confirmed starters in slots 1-6.
    - If lineups are not available yet, allows only stronger projected hitters.
    """
    cols = ["category","bet_type","playerName","teamName","game","opponent_pitcher","opponent_pitcher_team","opponent_pitcher_pick_type","opponent_pitcher_sample","lineup_status","batting_order_slot","starter_only_flag","HR_score","Hit_score","contact_quality_score","hit_quality_label","hit_pct_last_10","hit_pct_last_5","current_hit_streak","recent_cash_rate","recent_cash_record","recent_cash_sample","recent_cash_last_10","confidence","park_favorability","stack_tag","reason"]

    if player_rows is None or player_rows.empty or pitcher_metrics is None or pitcher_metrics.empty:
        return pd.DataFrame([{"category":"Info","bet_type":"No Plays","reason":"No refined picks met today’s filters"}], columns=cols)

    pm = pitcher_metrics[["teamName","opponentTeam","pitcherName","pick_type","sample_flag","short_leash_flag"]].drop_duplicates().rename(columns={
        "teamName":"opponent_pitcher_team",
        "opponentTeam":"teamName",
        "pitcherName":"opponent_pitcher",
        "pick_type":"opponent_pitcher_pick_type",
        "sample_flag":"opponent_pitcher_sample",
    })

    if game_rankings is not None and not game_rankings.empty:
        ctx_cols = [c for c in ["teamName","opponentTeam","game","offense_score","edge_vs_opponent","recommended_play","projected_win_pct","volatility_label","parlay_grade"] if c in game_rankings.columns]
        ctx = game_rankings[ctx_cols].drop_duplicates()
    else:
        ctx = pd.DataFrame(columns=["teamName","opponentTeam","game","offense_score","edge_vs_opponent","recommended_play"])

    rows = player_rows.merge(pm, on="teamName", how="left")
    merge_keys = ["teamName", "opponentTeam"] if "opponentTeam" in rows.columns and "opponentTeam" in ctx.columns else ["teamName"]
    rows = rows.merge(ctx, on=merge_keys, how="left")

    # Defensive defaults
    for c, default in [("Hit_score", 0), ("HR_score", 0), ("contact_quality_score", 0), ("hit_pct_last_10", 0), ("hit_pct_last_5", 0), ("current_hit_streak", 0), ("batting_order_slot", 99), ("offense_score", 0), ("edge_vs_opponent", 0)]:
        if c not in rows.columns:
            rows[c] = default
        rows[c] = pd.to_numeric(rows[c], errors="coerce").fillna(default)

    for c, default in [("lineup_status", "Unknown"), ("starter_only_flag", False), ("opponent_pitcher_pick_type", "Neutral"), ("park_favorability", "Unknown")]:
        if c not in rows.columns:
            rows[c] = default

    if "game" not in rows.columns:
        rows["game"] = None

    rows["game_str"] = rows["game"].astype(str)
    valid_game = rows["game_str"].notna() & (~rows["game_str"].str.strip().isin(["", "—", "-", "None", "nan"]))
    confirmed_count = int(rows["lineup_status"].astype(str).eq("Confirmed Starter").sum())
    use_confirmed_only = confirmed_count >= 5

    rows["slot_num"] = pd.to_numeric(rows["batting_order_slot"], errors="coerce").fillna(99)
    confirmed = rows["lineup_status"].astype(str).eq("Confirmed Starter")

    if use_confirmed_only:
        quality_mask = confirmed & (rows["slot_num"] <= 6) & (rows["Hit_score"] >= 3.85) & (rows["contact_quality_score"] >= 1.55) & (rows["hit_pct_last_10"] >= 60)
        mode_label = "confirmed-only"
    else:
        # Pregame mode: avoid flooding. Unknowns need a higher score because lineups are not locked yet.
        quality_mask = (
            (confirmed & (rows["Hit_score"] >= 3.75)) |
            ((~confirmed) & (rows["Hit_score"] >= 4.45) & (rows["contact_quality_score"] >= 1.85) & (rows["hit_pct_last_10"] >= 70))
        )
        mode_label = "pregame-projected"

    hit_pool = rows[
        valid_game &
        quality_mask &
        (~rows["opponent_pitcher_pick_type"].astype(str).eq("Strong SP"))
    ].copy()

    if hit_pool.empty:
        return pd.DataFrame([{"category":"Info","bet_type":"No Plays","reason":"No refined picks met Top-6 quality gate"}], columns=cols)

    # Rank by hitter quality first, then lineup quality, then game edge.
    # Rolling slate logic:
    # - max 2 refined picks per game, so late games can still appear on later refreshes
    # - max 2 per team, to avoid one offense flooding the refined card
    # - max 10 total refined picks, so this remains a true refined list rather than a research dump
    hit_pool["lineup_rank"] = hit_pool["lineup_status"].astype(str).eq("Confirmed Starter").astype(int)
    hit_pool = hit_pool.sort_values(["Hit_score", "contact_quality_score", "lineup_rank", "edge_vs_opponent", "slot_num"], ascending=[False, False, False, False, True])
    hit_pool = hit_pool.drop_duplicates(subset=["playerName", "teamName", "game"], keep="first")

    selected_rows = []
    team_counts = {}
    game_counts = {}
    for _, cand in hit_pool.iterrows():
        team = cand.get("teamName")
        game = cand.get("game")
        if team_counts.get(team, 0) >= 2:
            continue
        if game_counts.get(game, 0) >= 2:
            continue
        selected_rows.append(cand)
        team_counts[team] = team_counts.get(team, 0) + 1
        game_counts[game] = game_counts.get(game, 0) + 1
        if len(selected_rows) >= 10:
            break

    if selected_rows:
        hit_pool = pd.DataFrame(selected_rows)
    else:
        hit_pool = hit_pool.iloc[0:0].copy()

    picks = []
    for _, r in hit_pool.iterrows():
        score = nz(r.get("Hit_score"))
        picks.append({
            "category":"Hit Pick",
            "bet_type":"1+ Hit",
            "playerName":r.get("playerName"),
            "teamName":r.get("teamName"),
            "game":r.get("game"),
            "opponent_pitcher":r.get("opponent_pitcher"),
            "opponent_pitcher_team":r.get("opponent_pitcher_team"),
            "opponent_pitcher_pick_type":r.get("opponent_pitcher_pick_type"),
            "opponent_pitcher_sample":r.get("opponent_pitcher_sample"),
            "lineup_status":r.get("lineup_status"),
            "batting_order_slot":r.get("batting_order_slot"),
            "starter_only_flag":r.get("starter_only_flag"),
            "HR_score":r.get("HR_score"),
            "Hit_score":score,
            "contact_quality_score":r.get("contact_quality_score"),
            "hit_quality_label":r.get("hit_quality_label"),
            "hit_pct_last_10":r.get("hit_pct_last_10"),
            "hit_pct_last_5":r.get("hit_pct_last_5"),
            "current_hit_streak":r.get("current_hit_streak"),
            "recent_cash_rate":r.get("recent_cash_rate"),
            "recent_cash_record":r.get("recent_cash_record"),
            "recent_cash_sample":r.get("recent_cash_sample"),
            "recent_cash_last_10":r.get("recent_cash_last_10"),
            "confidence":r.get("confidence"),
            "park_favorability":r.get("park_favorability"),
            "stack_tag":"Rolling refined",
            "reason":f"Rolling refined max2/game max10 {mode_label}; Hit_score {score:.3f}; contact {r.get('contact_quality_score')}; L10 hit {r.get('hit_pct_last_10')}%; slot {r.get('batting_order_slot')}; opp {r.get('opponent_pitcher_pick_type')}; edge {r.get('edge_vs_opponent')}",
        })

    return pd.DataFrame(picks, columns=cols)

def build_pitcher_line_value(pitcher_metrics):
    cols = [
        "pitcherName", "teamName", "opponentTeam", "pick_type", "sample_flag",
        "innings_pitched", "strikeouts", "earned_runs", "hits_allowed", "walks",
        "pitcher_score_adj", "avg_ip_per_start", "avg_k_per_start", "k_per_inning",
        "last2_ip_avg", "last2_k_avg", "last2_pitch_avg", "k5plus_last10_pct", "k6plus_last10_pct", "k7plus_last10_pct", "short_leash_flag",
        "opp_team_k_rate", "opp_team_k_tendency", "opp_k_matchup_bonus", "own_bullpen_grade",
        "projected_k_floor", "projected_k_mid", "projected_k_ceiling", "max_playable_k_line",
        "k_value_tier", "recommended_k_action", "safest_pitching_play", "notes",
        "probable_starter_name", "starter_status",
    ]
    if pitcher_metrics is None or pitcher_metrics.empty:
        return pd.DataFrame(columns=cols)
    rows = []
    for _, r in pitcher_metrics.iterrows():
        ip = nz(r.get("innings_pitched"))
        ks = nz(r.get("strikeouts"))
        gs = max(nz(r.get("games_started"), 1), 1)
        avg_ip = round(ip / gs, 3) if gs else ip
        avg_k = round(ks / gs, 3) if gs else ks
        kpi = round(ks / ip, 3) if ip else 0
        recent_ip = nz(r.get("last2_ip_avg"), avg_ip)
        recent_k = nz(r.get("last2_k_avg"), avg_k)
        recent_pitches = nz(r.get("last2_pitch_avg"), 0)
        opp_k_bonus = nz(r.get("opp_k_matchup_bonus"), 0)
        proj_base = round((avg_k * 0.40) + (recent_k * 0.50) + (opp_k_bonus * 0.90), 3)
        mid = max(0, round(proj_base))
        floor = max(0, mid - 1)
        ceil = mid + 1
        short_leash_flag = str(r.get("short_leash_flag") or "Unknown")
        if short_leash_flag.startswith("Yes"):
            max_line, tier, action = "", "Pass", "Pass - short leash risk"
        elif recent_ip >= 5.4 and recent_k >= 7 and recent_pitches >= 90 and opp_k_bonus >= 0.35:
            max_line, tier, action = 6.5, "Hammer", "Bet over up to 6.5"
        elif recent_ip >= 5.0 and recent_k >= 6 and recent_pitches >= 85:
            max_line, tier, action = 5.5, "Strong", "Bet over up to 5.5"
        elif recent_ip >= 4.5 and recent_k >= 5 and recent_pitches >= 80 and opp_k_bonus >= 0:
            max_line, tier, action = 4.5, "Lean", "Only bet over at 4.5"
        else:
            max_line, tier, action = "", "Pass", "Pass"
        safest = "Avoid" if tier == "Pass" else ("Over Ks / Over outs / Under ER" if tier in ("Hammer","Strong") else "Over outs / Under ER")
        rows.append({
            "pitcherName":r["pitcherName"],"teamName":r["teamName"],"opponentTeam":r["opponentTeam"],"pick_type":r["pick_type"],"sample_flag":r["sample_flag"],
            "innings_pitched":ip,"strikeouts":ks,"earned_runs":r.get("earned_runs"),"hits_allowed":r.get("hits_allowed"),"walks":r.get("walks"),
            "pitcher_score_adj":r.get("pitcher_score_adj"),"avg_ip_per_start":avg_ip,"avg_k_per_start":avg_k,"k_per_inning":kpi,
            "last2_ip_avg":r.get("last2_ip_avg"),"last2_k_avg":r.get("last2_k_avg"),"last2_pitch_avg":r.get("last2_pitch_avg"),
            "k5plus_last10_pct":r.get("k5plus_last10_pct"),"k6plus_last10_pct":r.get("k6plus_last10_pct"),"k7plus_last10_pct":r.get("k7plus_last10_pct"),
            "short_leash_flag":short_leash_flag,
            "opp_team_k_rate":r.get("opp_team_k_rate"),"opp_team_k_tendency":r.get("opp_team_k_tendency"),"opp_k_matchup_bonus":opp_k_bonus,
            "own_bullpen_grade":r.get("own_bullpen_grade"),
            "projected_k_floor":floor,"projected_k_mid":mid,"projected_k_ceiling":ceil,"max_playable_k_line":max_line,
            "k_value_tier":tier,"recommended_k_action":action,"safest_pitching_play":safest,
            "notes":f"Starter locked; recent form + team K layer + bullpen support. oppK={r.get('opp_team_k_tendency')} pen={r.get('own_bullpen_grade')}",
            "probable_starter_name":r.get("probable_starter_name"),"starter_status":r.get("starter_status"),
        })
    if not rows:
        return pd.DataFrame(columns=cols)
    return pd.DataFrame(rows, columns=cols).sort_values(["projected_k_mid","pitcher_score_adj"], ascending=False).reset_index(drop=True)



# ------------------------------
# K PROP MARKET VALIDATION V2
# ------------------------------
def _k_first_existing_column(df: pd.DataFrame, candidates: list[str]):
    for c in candidates:
        if c in df.columns:
            return c
    return None


def _k_norm_text(v) -> str:
    if v is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(v).strip().lower())


def load_current_k_market_lines(target_date: str) -> pd.DataFrame:
    """
    Optional current sportsbook K-line file.

    This is intentionally strict: K props are blocked from the Final Card unless
    a CURRENT market line is supplied. This prevents stale 5.5 lines from being
    treated as live value when the real book is already 7.5.

    Accepted files in OUTPUT_DIR:
      - k_lines_YYYY-MM-DD.csv
      - posted_k_lines_YYYY-MM-DD.csv
      - current_k_lines_YYYY-MM-DD.csv
      - k_lines.csv
      - posted_k_lines.csv
      - current_k_lines.csv

    Accepted columns:
      pitcherName / pitcher / player / playerName / name
      teamName / team / pitcherTeam
      opponentTeam / opponent / opp
      current_k_line / posted_k_line / k_line / line / strikeout_line
      opening_k_line / open_k_line / opener_line   optional
      over_odds / k_over_odds / odds               optional
      book / sportsbook / source                   optional
    """
    configured = os.getenv("HR_K_LINES_CSV")
    candidates = []
    if configured:
        candidates.append(Path(configured))
    candidates.extend([
        OUTPUT_DIR / f"current_k_lines_{target_date}.csv",
        OUTPUT_DIR / f"posted_k_lines_{target_date}.csv",
        OUTPUT_DIR / f"k_lines_{target_date}.csv",
        OUTPUT_DIR / "current_k_lines.csv",
        OUTPUT_DIR / "posted_k_lines.csv",
        OUTPUT_DIR / "k_lines.csv",
    ])

    empty_cols = [
        "pitcherName", "teamName", "opponentTeam", "current_k_line",
        "opening_k_line", "k_over_odds", "k_line_book", "k_line_source",
        "_pitcher_key", "_team_key"
    ]

    path = next((p for p in candidates if p and p.exists()), None)
    if path is None:
        print_step("⚠️ No current K-line CSV found. K props will be blocked from Final Card until live lines are supplied.")
        return pd.DataFrame(columns=empty_cols)

    try:
        raw = pd.read_csv(path)
    except Exception as e:
        print_step(f"⚠️ Could not read K-line CSV {path}: {e}. K props will be blocked from Final Card.")
        return pd.DataFrame(columns=empty_cols)

    if raw is None or raw.empty:
        print_step(f"⚠️ K-line CSV {path} was empty. K props will be blocked from Final Card.")
        return pd.DataFrame(columns=empty_cols)

    pitcher_col = _k_first_existing_column(raw, ["pitcherName", "pitcher", "player", "playerName", "name"])
    team_col = _k_first_existing_column(raw, ["teamName", "team", "pitcherTeam"])
    opp_col = _k_first_existing_column(raw, ["opponentTeam", "opponent", "opp", "opponentName"])
    current_line_col = _k_first_existing_column(raw, ["current_k_line", "posted_k_line", "k_line", "line", "strikeout_line", "strikeouts_line", "so_line"])
    opening_line_col = _k_first_existing_column(raw, ["opening_k_line", "open_k_line", "opener_line", "opening_line"])
    odds_col = _k_first_existing_column(raw, ["over_odds", "k_over_odds", "odds", "price"])
    book_col = _k_first_existing_column(raw, ["book", "sportsbook", "source"])

    if pitcher_col is None or current_line_col is None:
        print_step(f"⚠️ K-line CSV {path} is missing pitcher/current line columns. K props will be blocked from Final Card.")
        return pd.DataFrame(columns=empty_cols)

    out = pd.DataFrame()
    out["pitcherName"] = raw[pitcher_col].astype(str)
    out["teamName"] = raw[team_col].astype(str) if team_col else ""
    out["opponentTeam"] = raw[opp_col].astype(str) if opp_col else ""
    out["current_k_line"] = pd.to_numeric(raw[current_line_col], errors="coerce")
    out["opening_k_line"] = pd.to_numeric(raw[opening_line_col], errors="coerce") if opening_line_col else None
    out["k_over_odds"] = raw[odds_col].astype(str) if odds_col else ""
    out["k_line_book"] = raw[book_col].astype(str) if book_col else "manual_csv"
    out["k_line_source"] = str(path)
    out = out[out["current_k_line"].notna()].copy()
    out["_pitcher_key"] = out["pitcherName"].apply(_k_norm_text)
    out["_team_key"] = out["teamName"].apply(_k_norm_text)
    out = out.drop_duplicates(subset=["_pitcher_key", "_team_key"], keep="last")
    print_step(f"✅ Loaded current K market lines from {path}: {len(out)} verified lines")
    return out


def apply_k_market_validation(pitcher_line_value: pd.DataFrame, target_date: str) -> pd.DataFrame:
    """
    Adds current-market validation for K props.

    Final Card rule:
    - Must have a supplied current sportsbook line
    - Current line must be <= model max playable line
    - Projected midpoint must clear current line by at least +1.25 Ks
    - If opening line is supplied and market moved up by >= 1.5 Ks, require +1.75 Ks edge
    """
    if pitcher_line_value is None:
        return pitcher_line_value

    df = pitcher_line_value.copy()
    if df.empty:
        return df

    df["_pitcher_key"] = df["pitcherName"].apply(_k_norm_text) if "pitcherName" in df.columns else ""
    df["_team_key"] = df["teamName"].apply(_k_norm_text) if "teamName" in df.columns else ""

    lines = load_current_k_market_lines(target_date)

    for c in ["current_k_line", "opening_k_line", "k_over_odds", "k_line_book", "k_line_source"]:
        if c not in df.columns:
            df[c] = None

    if lines.empty:
        df["k_market_line_status"] = "Missing current sportsbook K line - blocked from Final Card"
        df["k_market_ok"] = False
        df["k_edge_vs_market"] = None
        df["k_line_movement"] = None
        df["k_market_moved_up_flag"] = False
    else:
        df = df.merge(
            lines[["_pitcher_key", "_team_key", "current_k_line", "opening_k_line", "k_over_odds", "k_line_book", "k_line_source"]],
            on=["_pitcher_key", "_team_key"],
            how="left",
            suffixes=("", "_line")
        )

        for col in ["current_k_line", "opening_k_line", "k_over_odds", "k_line_book", "k_line_source"]:
            line_col = f"{col}_line"
            if line_col in df.columns:
                df[col] = df[line_col].combine_first(df[col])
                df = df.drop(columns=[line_col])

        # Pitcher-only fallback if team names do not match exactly.
        missing = df["current_k_line"].isna()
        if missing.any():
            pitcher_only = lines.drop_duplicates(subset=["_pitcher_key"], keep="last")[["_pitcher_key", "current_k_line", "opening_k_line", "k_over_odds", "k_line_book", "k_line_source"]]
            fallback = df.loc[missing, ["_pitcher_key"]].merge(pitcher_only, on="_pitcher_key", how="left")
            for col in ["current_k_line", "opening_k_line", "k_over_odds", "k_line_book", "k_line_source"]:
                df.loc[missing, col] = fallback[col].values

        projected_mid = pd.to_numeric(df.get("projected_k_mid"), errors="coerce")
        current_line = pd.to_numeric(df.get("current_k_line"), errors="coerce")
        opening_line = pd.to_numeric(df.get("opening_k_line"), errors="coerce")
        model_max = pd.to_numeric(df.get("max_playable_k_line"), errors="coerce")

        df["k_edge_vs_market"] = (projected_mid - current_line).round(2)
        df["k_line_movement"] = (current_line - opening_line).round(2)
        df["k_market_moved_up_flag"] = df["k_line_movement"].fillna(0) >= 1.5

        required_edge = df["k_market_moved_up_flag"].map(lambda moved: 1.75 if moved else 1.25)
        df["k_market_ok"] = (
            current_line.notna()
            & model_max.notna()
            & (current_line <= model_max)
            & (df["k_edge_vs_market"] >= required_edge)
        )

        def _status(row):
            line = row.get("current_k_line")
            if pd.isna(line):
                return "Missing current sportsbook K line - blocked from Final Card"
            edge = row.get("k_edge_vs_market")
            model_max_line = row.get("max_playable_k_line")
            moved = bool(row.get("k_market_moved_up_flag"))
            move = row.get("k_line_movement")
            if not row.get("k_market_ok"):
                reason = f"Current line {line:g}; projected mid {row.get('projected_k_mid')}; edge {edge}; model max {model_max_line}"
                if moved:
                    reason += f"; market moved up {move:g} Ks"
                return reason + " - blocked from Final Card"
            reason = f"Current line {line:g}; projected mid {row.get('projected_k_mid')}; edge +{edge}; model max {model_max_line}"
            if moved:
                reason += f"; market moved up {move:g} Ks but edge still qualifies"
            return reason + " - eligible"

        df["k_market_line_status"] = df.apply(_status, axis=1)

    df = df.drop(columns=[c for c in ["_pitcher_key", "_team_key"] if c in df.columns], errors="ignore")
    return df






# ------------------------------
# HR VALUE PROFILE ENGINE V1
# ------------------------------
def hr_value_bucket(home_runs, league_avg_hr=None):
    """Bucket HR totals into value-focused groups. Excludes 0-HR players from target logic.

    The HR Value Watch target is NOT low-power longshots. It is the near-average
    value band: players close to the dynamic league average, slightly below it,
    at it, or just above it. Someone way below average is excluded from the
    primary value bucket even if they are overdue.
    """
    hrs = nz(home_runs, 0)
    avg = nz(league_avg_hr, 0)
    try:
        hrs = float(hrs)
        avg = float(avg)
    except Exception:
        return "Unknown"
    if hrs <= 0:
        return "No HR profile"
    if hrs <= 2:
        return "Low Power"
    if avg:
        lower = max(3.0, avg - 2.0)
        upper = avg + 1.0
        if lower <= hrs <= upper:
            return "Near Avg HR Value Range"
        if hrs < lower:
            return "Below Value Band"
        if hrs <= avg + 4:
            return "Above Average Power"
    return "Elite / Priced-Up Power"


def add_hr_value_profile(player_rows: pd.DataFrame) -> pd.DataFrame:
    """Add dynamic HR value fields to player rows and HR_Drought output.

    Philosophy:
    - Exclude 0-HR players from the league average.
    - Prefer hitters with real power, but not obvious/priced-up elite power.
    - HR value target = 3+ HRs and at/below the dynamic league average plus a small cushion.
    - Environment still matters: bad parks and Strong SPs reduce the score.
    """
    if player_rows is None or player_rows.empty or "homeRuns" not in player_rows.columns:
        return player_rows

    df = player_rows.copy()
    hr_numeric = pd.to_numeric(df["homeRuns"], errors="coerce").fillna(0)
    positive = hr_numeric[hr_numeric > 0]
    league_avg_hr = round(float(positive.mean()), 2) if len(positive) else 0.0
    league_median_hr = round(float(positive.median()), 2) if len(positive) else 0.0

    df["league_avg_hr_excl_zero"] = league_avg_hr
    df["league_median_hr_excl_zero"] = league_median_hr
    df["hr_vs_league_avg"] = (hr_numeric - league_avg_hr).round(2)
    df["hr_value_bucket"] = hr_numeric.apply(lambda x: hr_value_bucket(x, league_avg_hr))

    # Dynamic sweet spot: close to league average, not way below it.
    # Example: if today's average is 6.44 HR, the value watch target is roughly 5-7 HR.
    # This avoids 3-HR/low-power names showing as Primary HR Value Targets when the league average is ~6+.
    hr_value_lower_bound = max(3.0, league_avg_hr - 2.0) if league_avg_hr else 3.0
    hr_value_upper_bound = (league_avg_hr + 1.0) if league_avg_hr else 99.0
    df["hr_value_lower_bound"] = round(hr_value_lower_bound, 2)
    df["hr_value_upper_bound"] = round(hr_value_upper_bound, 2)
    df["hr_value_band_distance"] = 0.0
    df.loc[hr_numeric < hr_value_lower_bound, "hr_value_band_distance"] = (hr_value_lower_bound - hr_numeric).round(2)
    df.loc[hr_numeric > hr_value_upper_bound, "hr_value_band_distance"] = (hr_numeric - hr_value_upper_bound).round(2)
    df["hr_value_target_flag"] = (hr_numeric >= hr_value_lower_bound) & (hr_numeric <= hr_value_upper_bound)

    drought = pd.to_numeric(df.get("current_games_without_hr"), errors="coerce").fillna(0)
    avg_gap = pd.to_numeric(df.get("avg_games_between_hrs"), errors="coerce").fillna(0)
    drought_over_avg = (drought - avg_gap).clip(lower=0)
    df["hr_drought_over_avg"] = drought_over_avg.round(2)

    park = df.get("park_favorability", pd.Series(["Neutral"] * len(df), index=df.index)).astype(str)
    opp_type = df.get("opponent_pitcher_pick_type", pd.Series(["Neutral"] * len(df), index=df.index)).astype(str)
    slot = pd.to_numeric(df.get("batting_order_slot"), errors="coerce").fillna(9)

    park_adj = park.map({"Favorable": 1.25, "Neutral": 0.55, "Unfavorable": -1.20}).fillna(0.25)
    pitcher_adj = opp_type.map({
        "Short Leash Risk": 1.15,
        "Low Sample": 0.95,
        "Attack With Hitters": 1.25,
        "Neutral": 0.35,
        "K Upside": -0.35,
        "Strong SP": -1.75,
    }).fillna(0.0)
    slot_adj = slot.apply(lambda s: max(0.0, 7 - float(s)) * 0.18 if s <= 9 else 0.0)
    value_band_adj = df["hr_value_target_flag"].astype(int) * 1.65
    low_power_penalty = (hr_numeric <= 2).astype(int) * -2.25
    below_value_band_penalty = (hr_numeric < hr_value_lower_bound).astype(int) * -2.00
    elite_price_penalty = (hr_numeric > (league_avg_hr + 4.0)).astype(int) * -1.10

    # Controlled score: value range + overdue timing + environment. Not raw HR leader ranking.
    contact_proxy = pd.to_numeric(df.get("hr_contact_proxy"), errors="coerce").fillna(0)

    df["hr_value_score"] = (
        value_band_adj
        + low_power_penalty
        + below_value_band_penalty
        + elite_price_penalty
        + drought_over_avg.clip(upper=18) * 0.16
        + park_adj
        + pitcher_adj
        + slot_adj
        + contact_proxy.clip(upper=2.5) * 0.35
    ).round(3)

    def _profile(row):
        if nz(row.get("homeRuns"), 0) <= 0:
            return "Exclude - no HR sample"
        if row.get("hr_value_target_flag") and row.get("hr_value_score", 0) >= 3.0:
            return "Primary HR Value Target"
        if row.get("hr_value_target_flag"):
            return "Secondary HR Value Watch"
        if row.get("hr_value_bucket") == "Below Value Band":
            return "Exclude - below HR value band"
        if row.get("hr_value_bucket") == "Elite / Priced-Up Power":
            return "Priced-Up Power / Upside Only"
        if row.get("hr_value_bucket") == "Low Power":
            return "Low-Power Longshot"
        return "HR Watch"

    df["hr_value_profile"] = df.apply(_profile, axis=1)
    df["hr_value_reason"] = df.apply(
        lambda r: (
            f"HRs {r.get('homeRuns')} vs dynamic avg {r.get('league_avg_hr_excl_zero')} "
            f"target band {r.get('hr_value_lower_bound')}-{r.get('hr_value_upper_bound')}; "
            f"bucket {r.get('hr_value_bucket')}; drought over avg {r.get('hr_drought_over_avg')}; "
            f"park {r.get('park_favorability')}; opp {r.get('opponent_pitcher_pick_type')}; "
            f"score {r.get('hr_value_score')}"
        ),
        axis=1,
    )
    return df


def build_hr_value_watch(player_rows: pd.DataFrame) -> pd.DataFrame:
    """Research tab for the exact HR value profile the user wants."""
    cols = [
        "playerName", "teamName", "homeRuns", "league_avg_hr_excl_zero", "hr_vs_league_avg",
        "hr_value_lower_bound", "hr_value_upper_bound", "hr_value_band_distance",
        "hr_value_bucket", "hr_value_target_flag", "hr_value_score", "hr_contact_proxy", "hr_value_profile",
        "avg_games_between_hrs", "current_games_without_hr", "hr_drought_over_avg", "hr_status",
        "last_hr_date", "gamesPlayed", "park_favorability", "opponent_pitcher", "opponent_pitcher_pick_type",
        "lineup_status", "batting_order_slot", "starter_only_flag", "hr_value_reason"
    ]
    if player_rows is None or player_rows.empty:
        return pd.DataFrame(columns=cols)
    df = player_rows.copy()
    if "hr_value_score" not in df.columns:
        df = add_hr_value_profile(df)
    if "opponent_pitcher" not in df.columns and "auto_pitcher_name" in df.columns:
        df["opponent_pitcher"] = df["auto_pitcher_name"]
    if "hr_status" not in df.columns and "status" in df.columns:
        df["hr_status"] = df["status"]
    pool = df[
        (pd.to_numeric(df.get("homeRuns"), errors="coerce").fillna(0) > 0)
        & (df.get("hr_value_target_flag", False).astype(bool))
    ].copy()
    # Do not fall back to low-power/way-below-average players. If nobody fits the
    # near-average value band, show an empty watch table rather than noisy longshots.
    pool = pool.sort_values(["hr_value_score", "hr_drought_over_avg", "HR_score"], ascending=[False, False, False])
    for c in cols:
        if c not in pool.columns:
            pool[c] = None
    return pool[cols].head(40).reset_index(drop=True)

def build_plus_money_prop_sheet(player_rows: pd.DataFrame, pitcher_line_value: pd.DataFrame, game_rankings: pd.DataFrame) -> pd.DataFrame:
    """Build a research-only prop finder for plus-money/alt props.

    This does NOT assume odds are available. It flags candidates that fit the same kind
    of profile as public hit sheets: 7/10+ recent cash rate, stable role, and +100-or-better
    markets to check manually.
    """
    cols = [
        "prop_type", "bet_type", "pick", "team", "opponent", "game", "confidence", "model_grade",
        "recent_cash_rate", "season_rate", "lineup_status", "batting_order_slot",
        "projected_edge_note", "market_check", "reason"
    ]
    rows = []

    gr = game_rankings.copy() if game_rankings is not None and not game_rankings.empty else pd.DataFrame()
    game_lookup = {}
    team_run_projection_lookup = {}
    if not gr.empty:
        run_cols = ["projected_team_runs", "team_projected_runs", "implied_team_total", "team_total", "projected_runs"]
        for _, r in gr.iterrows():
            team = r.get("teamName")
            opp = r.get("opponentTeam")
            game_lookup[(team, opp)] = r.get("game")
            # Strict RBI gate support: only accept true/projected team run fields when present.
            # If no projected run field exists, RBI props are blocked instead of guessed.
            for rc in run_cols:
                if rc in gr.columns:
                    try:
                        val = float(r.get(rc))
                        if pd.notna(val):
                            team_run_projection_lookup[(team, opp)] = val
                            team_run_projection_lookup[team] = val
                            break
                    except Exception:
                        pass

    if player_rows is not None and not player_rows.empty:
        pr = player_rows.copy()
        pr["slot_num"] = pd.to_numeric(pr.get("batting_order_slot"), errors="coerce").fillna(99)
        pr["confirmed"] = pr.get("lineup_status", "Unknown").astype(str).eq("Confirmed Starter") if "lineup_status" in pr.columns else False
        pr["opp_type"] = pr.get("opponent_pitcher_pick_type", "Neutral") if "opponent_pitcher_pick_type" in pr.columns else "Neutral"

        for _, r in pr.iterrows():
            team = r.get("teamName")
            opp = r.get("opponentTeam") or r.get("opponent_pitcher_team")
            game = r.get("game") or game_lookup.get((team, opp))
            slot = nz(r.get("slot_num"), 99)
            confirmed = bool(r.get("confirmed"))
            opp_type = str(r.get("opp_type") or "Neutral")
            hit_l10 = r.get("hit_pct_last_10")
            hit_l5 = r.get("hit_pct_last_5")
            season_hit = r.get("season_hit_pct")
            tb10 = r.get("tb2plus_last10_pct")
            tb5 = r.get("tb2plus_last5_pct")
            run10 = r.get("run_last10_pct")
            rbi10 = r.get("rbi_last10_pct")
            hit_score = nz(r.get("Hit_score"))
            hr_score = nz(r.get("HR_score"))
            name = r.get("playerName")
            lineup_ok = confirmed and slot <= 6
            avoid_sp = opp_type == "Strong SP"

            # 2+ Total Bases: best for hot hitters with power/contact and top/middle lineup role.
            if tb10 is not None and tb10 >= 70 and lineup_ok and not avoid_sp:
                conf = "A+" if tb10 >= 80 and (tb5 is None or tb5 >= 60) and (hr_score >= 3.8 or hit_score >= 4.2) else "A"
                rows.append({
                    "prop_type": "2+ Total Bases", "bet_type": "2+ Total Bases", "pick": name, "team": team, "opponent": opp, "game": game,
                    "confidence": conf, "model_grade": round(hit_score + max(hr_score, 0) * 0.20, 3),
                    "recent_cash_rate": tb10, "season_rate": season_hit,
                    "lineup_status": r.get("lineup_status"), "batting_order_slot": r.get("batting_order_slot"),
                    "projected_edge_note": "TB profile: recent 2+ TB rate plus contact/power support",
                    "market_check": "Only play if +100 or better, or if model edge beats posted price",
                    "reason": f"2+TB L10 {tb10}%; L5 {tb5}; Hit_score {hit_score:.3f}; HR_score {hr_score:.3f}; slot {slot}; opp {opp_type}",
                })

            # 1+ Run: top of order + strong hit form + team/game environment.
            if run10 is not None and run10 >= 70 and confirmed and slot <= 3 and not avoid_sp:
                rows.append({
                    "prop_type": "1+ Run", "bet_type": "1+ Run", "pick": name, "team": team, "opponent": opp, "game": game,
                    "confidence": "A" if run10 >= 80 else "B", "model_grade": round(hit_score + (4 - min(slot, 4)) * 0.25, 3),
                    "recent_cash_rate": run10, "season_rate": season_hit,
                    "lineup_status": r.get("lineup_status"), "batting_order_slot": r.get("batting_order_slot"),
                    "projected_edge_note": "Run prop profile: top-order PA volume + recent run cash rate",
                    "market_check": "Best used for plus-money or SGP leg; verify lineup and team total",
                    "reason": f"Run L10 {run10}%; slot {slot}; Hit_score {hit_score:.3f}; opp {opp_type}",
                })

            # 1+ RBI: HARD Option-B gate from results audit.
            # Require ALL:
            # - batting 1-5
            # - team projected runs >= 4.8
            # - opposing pitcher not Strong SP
            # - recent RBI cash >= 70%
            # - confirmed starter
            # If projected team runs are unavailable, reject instead of guessing.
            projected_team_runs = team_run_projection_lookup.get((team, opp), team_run_projection_lookup.get(team))
            try:
                projected_team_runs_num = float(projected_team_runs) if projected_team_runs is not None else None
            except Exception:
                projected_team_runs_num = None

            rbi_gate_ok = (
                rbi10 is not None
                and rbi10 >= 70
                and confirmed
                and slot <= 5
                and not avoid_sp
                and projected_team_runs_num is not None
                and projected_team_runs_num >= 4.8
            )
            if rbi_gate_ok:
                rows.append({
                    "prop_type": "1+ RBI", "bet_type": "1+ RBI", "pick": name, "team": team, "opponent": opp, "game": game,
                    "confidence": "A" if rbi10 >= 80 else "B", "model_grade": round(hit_score + hr_score * 0.15 + projected_team_runs_num * 0.10, 3),
                    "recent_cash_rate": rbi10, "season_rate": season_hit,
                    "lineup_status": r.get("lineup_status"), "batting_order_slot": r.get("batting_order_slot"),
                    "projected_edge_note": "RBI Option-B gate: top-5 bat, team runs >=4.8, non-Strong-SP, 70%+ recent RBI cash, confirmed starter",
                    "market_check": "Only play if price is +100 or better and runners-ahead context is strong",
                    "reason": f"RBI L10 {rbi10}%; slot {slot}; team projected runs {projected_team_runs_num}; Hit_score {hit_score:.3f}; HR_score {hr_score:.3f}; opp {opp_type}",
                })

    if pitcher_line_value is not None and not pitcher_line_value.empty:
        kdf = pitcher_line_value.copy()
        for _, r in kdf.iterrows():
            if str(r.get("short_leash_flag") or "").startswith("Yes"):
                continue
            k5 = r.get("k5plus_last10_pct")
            k6 = r.get("k6plus_last10_pct")
            mid = nz(r.get("projected_k_mid"))
            opp_tendency = r.get("opp_team_k_tendency")
            conf = None
            if k5 is not None and k5 >= 70 and mid >= 5:
                conf = "A+" if k5 >= 80 and mid >= 6 else "A"
                rows.append({
                    "prop_type": "5+ Strikeouts", "bet_type": "5+ Strikeouts", "pick": r.get("pitcherName"), "team": r.get("teamName"), "opponent": r.get("opponentTeam"),
                    "game": None, "confidence": conf, "model_grade": round(mid + nz(r.get("opp_k_matchup_bonus")), 3),
                    "recent_cash_rate": k5, "season_rate": r.get("avg_k_per_start"),
                    "lineup_status": "Confirmed Starter", "batting_order_slot": None,
                    "projected_edge_note": "Alt-K profile: recent 5+ K cash rate + projected K mid + opponent K tendency",
                    "market_check": "Only play if +100 or better; compare to current sportsbook alt-K price",
                    "reason": f"5+K L10 {k5}%; 6+K L10 {k6}; projected mid {mid}; oppK {opp_tendency}; tier {r.get('k_value_tier')}",
                })

    if not rows:
        return pd.DataFrame([{"prop_type":"Info", "bet_type":"Info", "pick":"No plus-money prop candidates", "reason":"No props met recent cash-rate and role filters"}], columns=cols)

    out = pd.DataFrame(rows)
    if "bet_type" not in out.columns:
        out["bet_type"] = out.get("prop_type")
    else:
        out["bet_type"] = out["bet_type"].fillna(out.get("prop_type"))
        out.loc[out["bet_type"].astype(str).str.strip().isin(["", "—", "None", "nan"]), "bet_type"] = out.get("prop_type")
    out["sort_rank"] = out["confidence"].map({"A+": 1, "A": 2, "B": 3}).fillna(9)
    out = out.sort_values(["sort_rank", "recent_cash_rate", "model_grade"], ascending=[True, False, False])
    out = out.drop(columns=["sort_rank"])
    return out.head(30).reset_index(drop=True)[cols]

def build_daily_card(game_rankings, refined_picks, pitcher_line_value, hr_drought):
    rows = []
    used_teams = set()

    ml = game_rankings[game_rankings["recommended_play"].isin(["Moneyline Lean","Stack Spot"])].sort_values("edge_vs_opponent", ascending=False)
    if not ml.empty:
        for _, best in ml.iterrows():
            if best["teamName"] in used_teams:
                continue
            rows.append({"section":"Best Overall","play_type":"Best Moneyline","pick":f"{best['teamName']} ML","team":best["teamName"],"opponent":best["opponentTeam"],"confidence":best["win_rating"],"why_it_made_the_card":f"Edge {best['edge_vs_opponent']}; {best['pitcher_pick_type']} vs {best['opponent_pitcher_pick_type']}; volatility pen {best['volatility_penalty_ml']}","source_tab":"Game_Rankings"})
            used_teams.add(best["teamName"])
            break
    else:
        rows.append({"section":"Best Overall","play_type":"Best Moneyline","pick":"No qualified ML play","confidence":"Pass","why_it_made_the_card":"No qualifying ML edge","source_tab":"Game_Rankings"})

    hit = refined_picks[refined_picks["category"].eq("Hit Pick")] if not refined_picks.empty and "category" in refined_picks.columns else pd.DataFrame()
    if not hit.empty:
        for _, best in hit.sort_values(["Hit_score","batting_order_slot"], ascending=[False, True]).iterrows():
            if best["teamName"] in used_teams:
                continue
            rows.append({"section":"Best Overall","play_type":"Best Hit","pick":best["playerName"],"team":best["teamName"],"opponent":best.get("opponent_pitcher_team"),"confidence":"Strong" if (best.get("batting_order_slot") or 99) <= 6 else "Lean","why_it_made_the_card":best.get("reason"),"source_tab":"Refined_Picks"})
            used_teams.add(best["teamName"])
            break
    else:
        rows.append({"section":"Best Overall","play_type":"Best Hit","pick":"No qualified hit play","confidence":"Pass","why_it_made_the_card":"Refined picks sheet was empty","source_tab":"Refined_Picks"})

    hr = refined_picks[refined_picks["category"].eq("HR Pick")] if not refined_picks.empty and "category" in refined_picks.columns else pd.DataFrame()
    if not hr.empty:
        for _, best in hr.sort_values(["HR_score","batting_order_slot"], ascending=[False, True]).iterrows():
            if best["teamName"] in used_teams:
                continue
            rows.append({"section":"Best Overall","play_type":"Best HR","pick":best["playerName"],"team":best["teamName"],"opponent":best.get("opponent_pitcher_team"),"confidence":"Strong","why_it_made_the_card":best.get("reason"),"source_tab":"Refined_Picks"})
            used_teams.add(best["teamName"])
            break
    else:
        rows.append({"section":"Best Overall","play_type":"Best HR","pick":"No qualified HR play","confidence":"Pass","why_it_made_the_card":"Refined picks sheet was empty","source_tab":"Refined_Picks"})

    kval = pitcher_line_value[pitcher_line_value["starter_status"].eq("Confirmed")] if not pitcher_line_value.empty and "starter_status" in pitcher_line_value.columns else pitcher_line_value
    if kval is not None and not kval.empty:
        for _, best in kval.sort_values(["projected_k_mid","pitcher_score_adj"], ascending=False).iterrows():
            if best["teamName"] in used_teams or str(best.get("short_leash_flag") or "").startswith("Yes"):
                continue
            rows.append({"section":"Best Overall","play_type":"Best K Prop","pick":best["pitcherName"],"team":best["teamName"],"opponent":best["opponentTeam"],"confidence":best["k_value_tier"],"why_it_made_the_card":f"{best['recommended_k_action']}; projected {best['projected_k_floor']}-{best['projected_k_ceiling']} Ks; max line {best['max_playable_k_line']}","source_tab":"Pitcher_Line_Value"})
            rows.append({"section":"Secondary","play_type":"Safest Pitching Play","pick":best["pitcherName"],"team":best["teamName"],"opponent":best["opponentTeam"],"confidence":best["k_value_tier"],"why_it_made_the_card":best["safest_pitching_play"],"source_tab":"Pitcher_Line_Value"})
            used_teams.add(best["teamName"])
            break
    else:
        rows.append({"section":"Best Overall","play_type":"Best K Prop","pick":"No qualified K play","confidence":"Pass","why_it_made_the_card":"No confirmed probable starter qualified","source_tab":"Pitcher_Line_Value"})

    if not hr_drought.empty:
        watch = hr_drought[hr_drought["status"].astype(str).str.contains("Overdue", na=False)].copy()
        if not watch.empty:
            watch["status_rank"] = watch["status"].astype(str).str.extract(r"\+(\d+)").fillna(0).astype(int)
            for _, best in watch.sort_values(["status_rank","homeRuns"], ascending=[False, False]).iterrows():
                if best["teamName"] in used_teams:
                    continue
                rows.append({"section":"Secondary","play_type":"Drought HR Watch","pick":best["playerName"],"team":best["teamName"],"confidence":best["status"],"why_it_made_the_card":f"{best['homeRuns']} HR; park={best['park_favorability']}; drought={best['current_games_without_hr']} games","source_tab":"HR_Drought"})
                break
    return pd.DataFrame(rows)

def build_final_card(player_rows, game_rankings, pitcher_line_value):
    """Official Final Card: elite-only, best-of-best plays.

    Philosophy:
    - Final Card is NOT a broad recommendation list.
    - No automatic Moneyline plays until ML performance improves.
    - HR plays stay in HR Value Watch / Top Picks because they are high variance.
    - K props can qualify only with verified current market edge and A+ profile.
    - 1+ Hit picks must pass very strict contact/recent-form/lineup gates.
    """
    cols = ["slot","bet_type","pick","team","opponent","confidence","why_it_made_the_card","source_tab","final_card_tier"]
    rows = []
    used_players = set()
    team_counts = {}

    def can_use_team(team, limit=1):
        return team_counts.get(team, 0) < limit

    def add_row(slot, bet_type, pick, team, opponent, confidence, why, source_tab, tier="Elite"):
        rows.append({
            "slot": slot,
            "bet_type": bet_type,
            "pick": pick,
            "team": team,
            "opponent": opponent,
            "confidence": confidence,
            "why_it_made_the_card": why,
            "source_tab": source_tab,
            "final_card_tier": tier,
        })
        team_counts[team] = team_counts.get(team, 0) + 1
        used_players.add((team, pick))

    # Moneylines are intentionally removed from the official Final Card for now.
    # Audit showed ML is roughly coin-flip compared with the hit engine.
    # ML still appears in Game Rankings for research, but does not take an official Final Card slot.

    if player_rows is not None and not player_rows.empty:
        base = player_rows.copy()
        if "auto_pitcher_name" in base.columns:
            base = base[base["auto_pitcher_name"].notna()].copy()

        for c, default in [
            ("opponent_pitcher_pick_type", "Neutral"),
            ("starter_only_flag", False),
            ("batting_order_slot", 99),
            ("Hit_score", 0),
            ("contact_quality_score", 0),
            ("hit_pct_last_10", 0),
            ("hit_pct_last_5", 0),
            ("current_hit_streak", 0),
            ("recent_cash_rate", None),
            ("park_favorability", "Neutral"),
            ("totalHits", 0),
        ]:
            if c not in base.columns:
                base[c] = default

        base["lineup_ok"] = base["starter_only_flag"].fillna(False)
        base["slot_num"] = pd.to_numeric(base.get("batting_order_slot"), errors="coerce").fillna(99)
        base["Hit_score_num"] = pd.to_numeric(base.get("Hit_score"), errors="coerce").fillna(0)
        base["contact_quality_num"] = pd.to_numeric(base.get("contact_quality_score"), errors="coerce").fillna(0)
        base["hit_l10_num"] = pd.to_numeric(base.get("hit_pct_last_10"), errors="coerce").fillna(0)
        base["hit_l5_num"] = pd.to_numeric(base.get("hit_pct_last_5"), errors="coerce").fillna(0)
        base["streak_num"] = pd.to_numeric(base.get("current_hit_streak"), errors="coerce").fillna(0)
        base["recent_cash_num"] = pd.to_numeric(base.get("recent_cash_rate"), errors="coerce")
        base["recent_cash_for_sort"] = base["recent_cash_num"].fillna(0)

        # Elite-only hit gate. This is intentionally tight after the results audit.
        # Final Card should be the best-of-best only:
        # confirmed starter, top-5 lineup slot, strong hit/contact score,
        # 80%+ last-10 hit form, 70%+ recent model cash rate, and no Strong SP.
        hit_pool = base[
            (base["lineup_ok"] == True) &
            (base["slot_num"] <= 5) &
            (base["Hit_score_num"] >= 5.00) &
            (base["contact_quality_num"] >= 3.40) &
            (base["hit_l10_num"] >= 80) &
            (base["recent_cash_num"].notna()) &
            (base["recent_cash_num"] >= 0.70) &
            (base["opponent_pitcher_pick_type"].fillna("Neutral") != "Strong SP")
        ].copy()

        if not hit_pool.empty:
            hit_pool["elite_sort_score"] = (
                hit_pool["Hit_score_num"] * 1.00
                + hit_pool["contact_quality_num"] * 0.35
                + (hit_pool["hit_l10_num"] / 100.0) * 0.80
                + hit_pool["recent_cash_for_sort"] * 0.60
                + hit_pool["streak_num"].clip(upper=5) * 0.08
                - hit_pool["slot_num"] * 0.05
            ).round(3)

            hit_pool = hit_pool.sort_values(
                ["elite_sort_score", "Hit_score_num", "contact_quality_num", "hit_l10_num", "slot_num"],
                ascending=[False, False, False, False, True]
            ).drop_duplicates(subset=["playerName", "teamName"], keep="first")

            max_final_hits = 3
            hit_added = 0
            for _, r in hit_pool.iterrows():
                if hit_added >= max_final_hits:
                    break
                if (r.get("teamName"), r.get("playerName")) in used_players or not can_use_team(r.get("teamName"), 1):
                    continue
                slot_label = f"Elite {hit_added + 1}"
                add_row(
                    slot_label,
                    "1+ Hit",
                    r.get("playerName"),
                    r.get("teamName"),
                    r.get("opponentTeam"),
                    "A+",
                    (
                        f"Elite hit gate; Hit_score {r.get('Hit_score_num'):.3f}; contact {r.get('contact_quality_num'):.2f}; "
                        f"L10 hit {r.get('hit_l10_num')}%; slot {int(r.get('slot_num'))}; "
                        f"recent cash {round(float(r.get('recent_cash_num') or 0)*100,1)}%; "
                        f"opp {r.get('opponent_pitcher_pick_type')}; park {r.get('park_favorability')}"
                    ),
                    "Elite_Final_Hit_Model",
                    "Elite Hit"
                )
                hit_added += 1

    # Optional K prop: only A+ alt-K / K edge with verified current market line.
    if pitcher_line_value is not None and not pitcher_line_value.empty and len(rows) < 3:
        kdf = pitcher_line_value.copy()
        for c, default in [
            ("starter_status", ""), ("short_leash_flag", ""), ("k_value_tier", ""),
            ("projected_k_mid", 0), ("pitcher_score_adj", 0), ("current_k_line", None),
            ("k_market_ok", False), ("k_edge_vs_market", None),
            ("k_market_line_status", "Missing current sportsbook K line - blocked from Final Card"),
            ("k_line_book", ""), ("k_over_odds", ""),
        ]:
            if c not in kdf.columns:
                kdf[c] = default
        kdf["k_edge_num"] = pd.to_numeric(kdf.get("k_edge_vs_market"), errors="coerce").fillna(0)
        kdf["projected_k_mid_num"] = pd.to_numeric(kdf.get("projected_k_mid"), errors="coerce").fillna(0)
        k_pool = kdf[
            (kdf["starter_status"] == "Confirmed") &
            (~kdf["short_leash_flag"].astype(str).str.startswith("Yes", na=False)) &
            (kdf["k_value_tier"].astype(str).eq("Hammer")) &
            (kdf["k_market_ok"].astype(bool) == True) &
            (kdf["k_edge_num"] >= 1.75) &
            (kdf["projected_k_mid_num"] >= 6)
        ].copy().sort_values(["k_edge_num", "projected_k_mid_num", "pitcher_score_adj"], ascending=[False, False, False])

        for _, r in k_pool.iterrows():
            if not can_use_team(r.get("teamName"), 1):
                continue
            add_row(
                f"Elite {len(rows)+1}",
                "K Prop",
                r.get("pitcherName"),
                r.get("teamName"),
                r.get("opponentTeam"),
                "A+",
                (
                    f"Elite K gate; current line {float(r.get('current_k_line')):g}; projected mid {r.get('projected_k_mid')}; "
                    f"edge vs market +{float(r.get('k_edge_num')):.2f}; {r.get('k_market_line_status')}; "
                    f"book/source {r.get('k_line_book') or 'manual_csv'}"
                ),
                "Elite_K_Model",
                "Elite K"
            )
            break

    if not rows:
        return pd.DataFrame([{
            "slot":"Info",
            "bet_type":"No Plays",
            "pick":"No elite Final Card plays qualified",
            "team":"",
            "opponent":"",
            "confidence":"Pass",
            "why_it_made_the_card":"Elite-only Final Card thresholds removed all plays. Check Refined Picks / Plus Money Props for research candidates.",
            "source_tab":"Final_Card",
            "final_card_tier":"No Play"
        }], columns=cols)

    return pd.DataFrame(rows, columns=cols)

def header_map(ws):
    return {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

def color_status_col(ws, header_name="status"):
    h = header_map(ws)
    c = h.get(header_name)
    if not c:
        return
    for r in range(2, ws.max_row + 1):
        v = str(ws.cell(row=r, column=c).value or "")
        if v.startswith("On Pace"):
            ws.cell(row=r, column=c).fill = GREEN
        elif v.startswith("Slightly Overdue"):
            ws.cell(row=r, column=c).fill = YELLOW
        elif v.startswith("Overdue"):
            ws.cell(row=r, column=c).fill = RED

def highlight_top_rows(ws, n=10):
    for r in range(2, min(ws.max_row, n + 1) + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).fill = GREEN

def main(season: int, target_date: str):
    OUTPUT_DIR.mkdir(exist_ok=True)
    print_step("🚀 V40.1 final-card rebuild started...")
    sched_ctx, schedule_rows = get_schedule_game_context(target_date)

    now_et = dt.datetime.now(ZoneInfo("America/New_York"))
    eligible_schedule_rows = filter_pregame_schedule_rows(schedule_rows, now_et=now_et, buffer_minutes=0)
    print_step(f"⏱️ Pregame-eligible games for Final Card: {len(eligible_schedule_rows)} of {len(schedule_rows)}")

    all_players = build_scheduled_player_pool(schedule_rows, season)
    lineup_map, slot_map = get_confirmed_lineups(target_date)
    locked_players = build_locked_player_pool(all_players, lineup_map, slot_map)

    print_step("🧠 Building full-slate player pool (all scheduled players) and tagging confirmed lineups when available ...")
    scoring_pool = all_players.copy()
    if scoring_pool.empty:
        player_rows = pd.DataFrame()
    else:
        scoring_pool["lineup_status"] = scoring_pool.apply(
            lambda r: lineup_map.get((r["teamName"], normalize_name(r["playerName"])), "Unknown"), axis=1
        )
        scoring_pool["batting_order_slot"] = scoring_pool.apply(
            lambda r: slot_map.get((r["teamName"], normalize_name(r["playerName"])), None), axis=1
        )
        scoring_pool["starter_only_flag"] = scoring_pool["lineup_status"].eq("Confirmed Starter")
        player_rows = build_hit_hr_rows(scoring_pool, season, sched_ctx)
    player_rows = player_rows[player_rows["auto_pitcher_name"].notna()].copy() if not player_rows.empty else player_rows

    pitcher_metrics = build_pitcher_metrics(schedule_rows, season)
    team_context_df = build_team_context_df(schedule_rows, season)
    player_rows = enrich_player_rows_with_team_context(player_rows, pitcher_metrics, team_context_df)
    player_rows = add_contact_quality_engine(player_rows)
    pitcher_metrics = enrich_pitcher_metrics_with_team_context(pitcher_metrics, team_context_df)
    pitcher_line_value = build_pitcher_line_value(pitcher_metrics)
    pitcher_line_value = apply_k_market_validation(pitcher_line_value, target_date)
    game_rankings = build_game_rankings(schedule_rows, player_rows, player_rows, pitcher_metrics)
    game_rankings = add_game_edge_engine(game_rankings)
    print_step("✅ Game Edge Engine applied to Game_Rankings")

    # Final Card protection: only generate actionable picks from games that have NOT started.
    # Games/Research still show the full slate for context.
    eligible_games = set(
        (eligible_schedule_rows.get("away_team", pd.Series(dtype=object)).fillna("")
         + " @ "
         + eligible_schedule_rows.get("home_team", pd.Series(dtype=object)).fillna("")).tolist()
    )
    eligible_teams = set(
        eligible_schedule_rows.get("away_team", pd.Series(dtype=object)).dropna().tolist()
        + eligible_schedule_rows.get("home_team", pd.Series(dtype=object)).dropna().tolist()
    )

    pregame_player_rows = player_rows[player_rows["teamName"].isin(eligible_teams)].copy() if not player_rows.empty else player_rows
    pregame_pitcher_line_value = pitcher_line_value[pitcher_line_value["teamName"].isin(eligible_teams)].copy() if not pitcher_line_value.empty else pitcher_line_value
    pregame_game_rankings = game_rankings[game_rankings["game"].isin(eligible_games)].copy() if not game_rankings.empty else game_rankings

    refined_picks = build_refined_picks(pregame_player_rows, pitcher_metrics, pregame_game_rankings)

    opp_map = pitcher_metrics[["opponentTeam","pitcherName","pick_type"]].drop_duplicates().rename(columns={
        "opponentTeam":"teamName","pitcherName":"opponent_pitcher","pick_type":"opponent_pitcher_pick_type"
    }) if not pitcher_metrics.empty else pd.DataFrame(columns=["teamName","opponent_pitcher","opponent_pitcher_pick_type"])

    # Ensure Final_Card logic always has opponent pitcher type available on player rows.
    if not player_rows.empty:
        if "opponent_pitcher_pick_type" not in player_rows.columns:
            player_rows = player_rows.merge(opp_map[["teamName", "opponent_pitcher_pick_type"]].drop_duplicates(), on="teamName", how="left")
        else:
            missing_mask = player_rows["opponent_pitcher_pick_type"].isna()
            if missing_mask.any():
                fill_map = opp_map[["teamName", "opponent_pitcher_pick_type"]].drop_duplicates()
                player_rows = player_rows.merge(fill_map, on="teamName", how="left", suffixes=("", "_fill"))
                player_rows["opponent_pitcher_pick_type"] = player_rows["opponent_pitcher_pick_type"].fillna(player_rows["opponent_pitcher_pick_type_fill"])
                player_rows = player_rows.drop(columns=["opponent_pitcher_pick_type_fill"])

    # Dynamic HR Value Profile: league average recalculates every refresh, excluding 0-HR hitters.
    player_rows = add_hr_value_profile(player_rows)
    hr_value_watch = build_hr_value_watch(player_rows)

    hr_drought_cols = [c for c in ["season","teamName","playerName","avg_games_between_hrs","current_games_without_hr","longest_games_without_hr","hr_status","homeRuns","league_avg_hr_excl_zero","league_median_hr_excl_zero","hr_vs_league_avg","hr_value_bucket","hr_value_target_flag","hr_value_score","hr_value_profile","hr_drought_over_avg","hr_value_reason","last_hr_date","gamesPlayed","park_favorability","lineup_status","batting_order_slot","starter_only_flag"] if c in player_rows.columns]
    hr_drought = player_rows[hr_drought_cols].rename(columns={"hr_status":"status"}).merge(opp_map, on="teamName", how="left")
    hit_drought = player_rows[["season","teamName","playerName","avg_games_between_hits","current_games_without_hit","longestHitDrought","hit_status","totalHits","gamesPlayed","park_favorability","lineup_status","batting_order_slot","starter_only_flag"]].rename(columns={"hit_status":"status"}).merge(opp_map, on="teamName", how="left")

    daily_card = build_daily_card(pregame_game_rankings, refined_picks, pregame_pitcher_line_value, hr_drought)
    final_card = build_final_card(pregame_player_rows, pregame_game_rankings, pregame_pitcher_line_value)
    plus_money_props = build_plus_money_prop_sheet(pregame_player_rows, pregame_pitcher_line_value, pregame_game_rankings)

    # Recent Cash Rate layer: adds player/bet-type history to the research tables.
    # This is read-only and will not reset or overwrite result history.
    recent_cash_history_rows = _load_recent_cash_history()
    refined_picks = add_recent_cash_rate_columns(refined_picks, recent_cash_history_rows, window=10)
    refined_picks = assign_refined_pick_confidence(refined_picks)
    plus_money_props = add_recent_cash_rate_columns(plus_money_props, recent_cash_history_rows, window=10)

    ts = dt.datetime.now().strftime("%Y-%m-%d_%H%M")
    outfile = OUTPUT_DIR / f"HR_Hit_Drought_v40_stats-{season}_{ts}.xlsx"

    print_step("💾 Writing workbook ...")
    with pd.ExcelWriter(outfile, engine="openpyxl") as writer:
        pd.DataFrame([
            ("requested_season", season),
            ("target_game_date", target_date),
            ("message", "v40 Value Top Picks: dynamic HR value profile + plus-money props + hit quality engine"),
            ("locked_players_count", len(locked_players)),
            ("pregame_eligible_games_for_final_card", len(eligible_schedule_rows)),
            ("run_time_et", now_et.strftime("%Y-%m-%d %I:%M %p ET").replace(" 0", " ")),
        ], columns=["field","value"]).to_excel(writer, sheet_name="Run_Info", index=False)

        schedule_rows.to_excel(writer, sheet_name="Schedule_Context", index=False)
        team_context_df.to_excel(writer, sheet_name="Team_Context", index=False)
        hr_drought.to_excel(writer, sheet_name="HR_Drought", index=False)
        hit_drought.to_excel(writer, sheet_name="Hit_Drought", index=False)
        pitcher_metrics.to_excel(writer, sheet_name="Pitcher_Metrics", index=False)
        pitcher_line_value.to_excel(writer, sheet_name="Pitcher_Line_Value", index=False)
        game_rankings.to_excel(writer, sheet_name="Game_Rankings", index=False)
        refined_picks.to_excel(writer, sheet_name="Refined_Picks", index=False)
        daily_card.to_excel(writer, sheet_name="Daily_Card", index=False)
        final_card.to_excel(writer, sheet_name="Final_Card", index=False)
        plus_money_props.to_excel(writer, sheet_name="Plus_Money_Props", index=False)
        hr_value_watch.to_excel(writer, sheet_name="HR_Value_Watch", index=False)

        top_hr = pd.DataFrame()
        top_hit = pd.DataFrame()
        if not player_rows.empty:
            # Top Picks should be value probability spots, not just raw HR/stat leaders.
            hr_pool = player_rows.copy()
            if "hr_value_score" not in hr_pool.columns:
                hr_pool = add_hr_value_profile(hr_pool)
            hr_pool = hr_pool[pd.to_numeric(hr_pool.get("homeRuns"), errors="coerce").fillna(0) > 0].copy()
            preferred_hr = hr_pool[hr_pool.get("hr_value_target_flag", False).astype(bool)].copy() if "hr_value_target_flag" in hr_pool.columns else hr_pool.iloc[0:0].copy()
            if preferred_hr.empty:
                preferred_hr = hr_pool.copy()
            hr_cols = [c for c in ["playerName","teamName","auto_pitcher_name","auto_pitcher_hand","homeRuns","league_avg_hr_excl_zero","hr_vs_league_avg","hr_value_bucket","hr_value_target_flag","hr_value_score","hr_contact_proxy","hr_value_profile","HR_score","current_games_without_hr","avg_games_between_hrs","hr_drought_over_avg","park_favorability","opponent_pitcher_pick_type","batting_order_slot","lineup_status","starter_only_flag","hr_value_reason"] if c in preferred_hr.columns]
            top_hr = preferred_hr.sort_values(["hr_value_score","hr_drought_over_avg","HR_score"], ascending=[False, False, False]).head(10)[hr_cols].copy()
            top_hr.insert(0, "type", "HR")

            hit_cols = [c for c in ["playerName","teamName","auto_pitcher_name","auto_pitcher_hand","Hit_score","contact_quality_score","hit_quality_label","hit_pct_last_10","hit_pct_last_5","current_hit_streak","contact_momentum_bonus","batting_order_slot","lineup_status","starter_only_flag"] if c in player_rows.columns]
            top_hit = player_rows.nlargest(10, "Hit_score")[hit_cols].copy()
            top_hit.insert(0, "type", "HIT")
        top_picks = pd.concat([top_hr, top_hit], ignore_index=True)
        top_picks = add_recent_cash_rate_columns(top_picks, recent_cash_history_rows, window=10)
        top_picks.to_excel(writer, sheet_name="Top_Picks", index=False)

    wb = load_workbook(outfile)
    for s in ["HR_Drought","Hit_Drought"]:
        if s in wb.sheetnames:
            color_status_col(wb[s], "status")
    for s in ["Pitcher_Metrics","Pitcher_Line_Value","Game_Rankings","Daily_Card","Final_Card","Top_Picks","Team_Context","Plus_Money_Props","HR_Value_Watch"]:
        if s in wb.sheetnames:
            highlight_top_rows(wb[s], 10)
    wb.save(outfile)

    json_filename = f"HR_Hit_Drought_v40_appdata-{season}_{target_date}_{ts}.json"
    json_output_path = OUTPUT_DIR / json_filename

    app_payload = build_app_payload(
        target_date=target_date,
        final_card_df=final_card,
        player_rows=player_rows,
        game_rankings=game_rankings,
        pitcher_metrics=pitcher_metrics,
        pitcher_line_value=pitcher_line_value,
        hr_drought=hr_drought,
        hit_drought=hit_drought,
        top_picks=top_picks,
        refined_picks=refined_picks,
        plus_money_props=plus_money_props,
        hr_value_watch=hr_value_watch
    )

    save_app_json(app_payload, json_output_path)
    print_step(f"🧾 JSON created: {json_output_path}")

    print_step("✅ DONE!")
    print_step(f"Created: {outfile}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="v40 final-card rebuild")
    parser.add_argument("--season", type=int, default=DEFAULT_SEASON)
    parser.add_argument("--date", type=str, default=dt.date.today().strftime("%Y-%m-%d"), help="Game date in YYYY-MM-DD format. Defaults to today if omitted.")
    args = parser.parse_args()
    main(args.season, args.date)
